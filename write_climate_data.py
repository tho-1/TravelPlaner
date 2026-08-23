"""
Helper script to write monthly climate data into Destinations.xlsx.

Monthly columns added (60 total):
  Jan High (C) .. Dec High (C)      - avg daily high temperature
  Jan Low (C)  .. Dec Low (C)       - avg daily low temperature (optional)
  Jan Rainy Days .. Dec Rainy Days  - avg rainy days per month
  Jan Rain (mm) .. Dec Rain (mm)    - avg rainfall in mm
  Jan AQI      .. Dec AQI           - avg Air Quality Index

Usage:
    from write_climate_data import write_monthly, print_progress

    write_monthly("Bogotá", {
        "Jan High (C)": 19, "Feb High (C)": 19, ...
        "Jan Low (C)": 7, ...
        "Jan Rainy Days": 9, ...
        "Jan Rain (mm)": 30, ...
        "Jan AQI": 55, ...
    })
"""
import openpyxl
from pathlib import Path

WORKBOOK = Path(__file__).resolve().parent / "Destinations.xlsx"

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# Annual summary columns (already written for first 15 destinations — keep as-is)
ANNUAL_COLS = [
    "Avg High Temp (°C)",
    "Avg Low Temp (°C)",
    "Avg Rainy Days/Month",
    "Avg Rain (mm/Month)",
    "Avg AQI",
]

# Monthly columns — 60 total
MONTHLY_COLS = (
    [f"{m} High (C)" for m in MONTHS] +
    [f"{m} Low (C)"  for m in MONTHS] +
    [f"{m} Rainy Days" for m in MONTHS] +
    [f"{m} Rain (mm)" for m in MONTHS] +
    [f"{m} AQI"      for m in MONTHS]
)

ALL_CLIMATE_COLS = ANNUAL_COLS + MONTHLY_COLS

# 2025 monthly aggregates fetched from Open-Meteo for Bogotá (4.7110, -74.0721).
# High/low are means of daily maxima/minima; rain is monthly precipitation sum;
# rainy days count days with at least 0.1 mm; AQI is the mean of hourly US AQI.
BOGOTA_CLIMATE = {
    "Avg High Temp (°C)": 19.6,
    "Avg Low Temp (°C)": 9.5,
    "Avg Rainy Days/Month": 29.1,
    "Avg Rain (mm/Month)": 128.6,
    "Avg AQI": 66.5,
    "Jan High (C)": 21.0, "Feb High (C)": 20.5, "Mar High (C)": 19.8,
    "Apr High (C)": 20.1, "May High (C)": 18.9, "Jun High (C)": 18.5,
    "Jul High (C)": 18.1, "Aug High (C)": 19.0, "Sep High (C)": 19.4,
    "Oct High (C)": 19.4, "Nov High (C)": 20.7, "Dec High (C)": 20.2,
    "Jan Low (C)": 8.1, "Feb Low (C)": 10.3, "Mar Low (C)": 9.8,
    "Apr Low (C)": 10.1, "May Low (C)": 10.1, "Jun Low (C)": 9.3,
    "Jul Low (C)": 9.5, "Aug Low (C)": 8.8, "Sep Low (C)": 8.9,
    "Oct Low (C)": 9.7, "Nov Low (C)": 10.0, "Dec Low (C)": 9.5,
    "Jan Rainy Days": 27, "Feb Rainy Days": 28, "Mar Rainy Days": 30,
    "Apr Rainy Days": 28, "May Rainy Days": 31, "Jun Rainy Days": 29,
    "Jul Rainy Days": 31, "Aug Rainy Days": 29, "Sep Rainy Days": 28,
    "Oct Rainy Days": 31, "Nov Rainy Days": 29, "Dec Rainy Days": 28,
    "Jan Rain (mm)": 92.3, "Feb Rain (mm)": 200.4, "Mar Rain (mm)": 256.6,
    "Apr Rain (mm)": 223.8, "May Rain (mm)": 127.3, "Jun Rain (mm)": 161.9,
    "Jul Rain (mm)": 60.4, "Aug Rain (mm)": 64.8, "Sep Rain (mm)": 71.0,
    "Oct Rain (mm)": 95.4, "Nov Rain (mm)": 124.0, "Dec Rain (mm)": 65.8,
    "Jan AQI": 70.5, "Feb AQI": 79.2, "Mar AQI": 84.0,
    "Apr AQI": 69.5, "May AQI": 62.7, "Jun AQI": 54.8,
    "Jul AQI": 49.9, "Aug AQI": 57.7, "Sep AQI": 61.5,
    "Oct AQI": 67.0, "Nov AQI": 68.3, "Dec AQI": 73.6,
}

# Monthly means of hourly US AQI for 2025, fetched from the Open-Meteo
# historical air-quality endpoint using each city-center coordinate. The
# endpoint reports CAMS global model/reanalysis estimates outside Europe,
# rather than direct local monitoring-station observations.
LAST_SIX_AQI = {
    "Panama City": [39.5, 45.2, 50.3, 44.0, 43.5, 40.1, 43.8, 43.3, 36.5, 30.5, 33.8, 38.0],
    "San José": [31.5, 38.0, 46.9, 42.4, 46.3, 46.5, 43.1, 51.4, 50.6, 48.3, 34.9, 38.0],
    "Yogyakarta": [138.3, 133.9, 147.5, 114.1, 128.8, 133.5, 112.8, 120.0, 119.3, 129.8, 138.9, 131.1],
    "Bishkek": [72.7, 71.3, 65.1, 55.7, 50.4, 60.7, 62.8, 55.2, 55.4, 60.5, 67.7, 68.7],
    "Dubai": [89.4, 100.8, 99.1, 113.1, 114.8, 128.5, 149.5, 149.0, 128.2, 110.6, 109.7, 105.5],
    "Montevideo": [36.7, 42.5, 41.0, 33.9, 38.2, 45.4, 54.3, 42.9, 42.6, 38.4, 41.2, 47.9],
}


def preserved_destination_climate(destination: str) -> dict:
    """Build writer fields from the preserved six-destination climate script."""
    from populate_destination_details import CLIMATE_DATA

    climate = CLIMATE_DATA[destination]
    data = {}
    for month, high, low, rainy_days, rain, aqi in zip(
        MONTHS,
        climate["high"],
        climate["low"],
        climate["rainy_days"],
        climate["rain"],
        LAST_SIX_AQI[destination],
    ):
        data[f"{month} High (C)"] = high
        data[f"{month} Low (C)"] = low
        data[f"{month} Rainy Days"] = rainy_days
        data[f"{month} Rain (mm)"] = rain
        data[f"{month} AQI"] = aqi

    data["Avg High Temp (°C)"] = round(sum(climate["high"]) / 12, 1)
    data["Avg Low Temp (°C)"] = round(sum(climate["low"]) / 12, 1)
    data["Avg Rainy Days/Month"] = round(sum(climate["rainy_days"]) / 12, 1)
    data["Avg Rain (mm/Month)"] = round(sum(climate["rain"]) / 12, 1)
    data["Avg AQI"] = round(sum(LAST_SIX_AQI[destination]) / 12, 1)
    return data


def _get_or_create_headers(ws) -> dict:
    """Return {header_name: 1-based column index}, adding missing climate columns."""
    headers = {cell.value: cell.column for cell in ws[1] if cell.value is not None}
    next_col = ws.max_column + 1

    for col_name in ALL_CLIMATE_COLS:
        if col_name not in headers:
            ws.cell(row=1, column=next_col, value=col_name)
            headers[col_name] = next_col
            next_col += 1
    return headers


def _find_dest_row(ws, destination: str, dest_col_idx: int):
    """Return the row index for the given destination name."""
    for row in ws.iter_rows(min_row=2):
        if row[dest_col_idx - 1].value == destination:
            return row[0].row
    return None


def write_monthly_row(row_idx: int, data: dict, overwrite: bool = False):
    """
    Write monthly (or annual) climate data for a specific 1-based row index.
    """
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb.active
    headers = _get_or_create_headers(ws)

    written = 0
    skipped = 0
    for col_name, value in data.items():
        col_idx = headers.get(col_name)
        if col_idx is None:
            print(f"[WARN] Column not found: {col_name!r}")
            continue
        cell = ws.cell(row=row_idx, column=col_idx)
        if not overwrite and cell.value is not None:
            skipped += 1
            continue
        cell.value = value
        written += 1

    wb.save(WORKBOOK)
    wb.close()
    dest = ws.cell(row=row_idx, column=1).value
    print(f"[OK] Row {row_idx} ({dest}): wrote {written} cells" +
          (f", skipped {skipped} already filled" if skipped else ""))


def write_monthly(destination: str, data: dict, overwrite: bool = False):
    """
    Write monthly (or annual) climate data for one destination.

    Args:
        destination: exact name as in spreadsheet col A
        data: dict mapping column names (from MONTHLY_COLS or ANNUAL_COLS) to values
        overwrite: if False, skip cells that already have a value
    """
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb.active
    headers = _get_or_create_headers(ws)

    dest_col_idx = headers.get("Destination")
    if dest_col_idx is None:
        dest_col_idx = 1

    # find all matching rows
    matched_rows = []
    for row in ws.iter_rows(min_row=2):
        if row[dest_col_idx - 1].value == destination:
            matched_rows.append(row[0].row)

    if not matched_rows:
        wb.close()
        print(f"[WARN] Destination not found in spreadsheet: {destination!r}")
        return

    written_total = 0
    skipped_total = 0
    for row_idx in matched_rows:
        written = 0
        skipped = 0
        for col_name, value in data.items():
            col_idx = headers.get(col_name)
            if col_idx is None:
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            if not overwrite and cell.value is not None:
                skipped += 1
                continue
            cell.value = value
            written += 1
        written_total += written
        skipped_total += skipped

    wb.save(WORKBOOK)
    wb.close()
    print(f"[OK] {destination} ({len(matched_rows)} rows): wrote {written_total} cells" +
          (f", skipped {skipped_total} already filled" if skipped_total else ""))



def print_progress():
    """Print fill status for every destination."""
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb.active
    headers = {cell.value: (cell.column - 1)  # 0-indexed for values tuple
               for cell in ws[1] if cell.value is not None}

    monthly_indices = [headers.get(c) for c in MONTHLY_COLS]
    total = len(MONTHLY_COLS)

    print(f"{'Destination':<30} {'Monthly':>10}  {'Annual':>8}")
    print("-" * 55)
    for row in ws.iter_rows(min_row=2, values_only=True):
        dest = row[0]
        if not dest:
            continue
        monthly_filled = sum(
            1 for idx in monthly_indices
            if idx is not None and idx < len(row) and row[idx] is not None
        )
        annual_indices = [headers.get(c) for c in ANNUAL_COLS]
        annual_filled = sum(
            1 for idx in annual_indices
            if idx is not None and idx < len(row) and row[idx] is not None
        )
        pct = int(100 * monthly_filled / total) if total else 0
        print(f"{str(dest):<30} {monthly_filled:>4}/{total}  ({pct:3}%)   {annual_filled}/{len(ANNUAL_COLS)} annual")

    wb.close()


if __name__ == "__main__":
    write_monthly("Bogotá", BOGOTA_CLIMATE, overwrite=True)
    for destination in LAST_SIX_AQI:
        write_monthly(destination, preserved_destination_climate(destination), overwrite=True)
    print_progress()
