from pathlib import Path

import openpyxl


WORKBOOK = Path(__file__).resolve().parent / "Destinations.xlsx"

DESTINATION_DETAILS = {
    "Panama City": {
        "selection": "x",
        "ideal": "December-April",
        "avoid": "May-November (heaviest rain)",
        "population": 1_500_000,
        "reachable": "Tocumen International Airport (PTY)",
        "visa": "no",
        "visa_requirement": "Visa-free for up to 90 days",
        "flight": 11.5,
        "profile": {
            "Why to Go There": "See the Panama Canal in operation, explore the UNESCO-listed Casco Antiguo, and combine a modern skyline with tropical rainforest and Pacific waterfronts.",
            "What to Expect": "A humid, fast-growing canal city with colonial streets, dense traffic, strong seafood and Afro-Caribbean influences, and easy access to nature.",
            "Recommended Stay": "3-4 days",
            "Avg. Cost/Day (3* Hotel & Food)": 75.0,
            "Rent a Car?": "No (Metro, taxis, and ride-hailing are easier in the city)",
            "Highlights": "Watch ships pass through the Miraflores Locks, walk Plaza de Francia and Las Bóvedas in Casco Antiguo, visit Panamá Viejo and the Biomuseo, and hike or birdwatch in Metropolitan Natural Park.",
            "Introduction Sentence": "Panama City pairs the engineering spectacle of the Panama Canal with a UNESCO-listed old quarter, tropical forest, and a lively Pacific-facing skyline.",
        },
    },
    "San José": {
        "selection": "no",
        "ideal": "December-April",
        "avoid": "September-October (very wet)",
        "population": 1_400_000,
        "reachable": "Juan Santamaria International Airport (SJO)",
        "visa": "no",
        "visa_requirement": "Visa-free for up to 90 days",
        "flight": 12.0,
        "profile": {
            "Why to Go There": "Use Costa Rica's capital as a cultural base for museums, coffee culture, markets, and easy trips into the Central Valley's volcano and cloud-forest landscapes.",
            "What to Expect": "A highland capital at about 1,030 metres, with mild temperatures, busy streets, strong public bus connections, and a concentrated museum and theatre district.",
            "Recommended Stay": "2-3 days",
            "Avg. Cost/Day (3* Hotel & Food)": 65.0,
            "Rent a Car?": "No (traffic is heavy; use buses, taxis, and ride-hailing)",
            "Highlights": "Visit the National Theater, National Museum, Pre-Columbian Gold Museum, and Jade Museum; eat at the Central Market; and relax in La Sabana Metropolitan Park.",
            "Introduction Sentence": "San José is Costa Rica's highland capital and cultural hub, combining coffee-country history, major museums, theatres, and a practical gateway to the Central Valley.",
        },
    },
    "Yogyakarta": {
        "selection": "no",
        "ideal": "June-September",
        "avoid": "December-February (wettest period)",
        "population": 4_000_000,
        "reachable": "Yogyakarta International Airport (YIA)",
        "visa": "no",
        "visa_requirement": "Visa on arrival or e-VOA",
        "flight": 14.0,
        "profile": {
            "Why to Go There": "Experience Java's strongest living arts tradition, royal court culture, batik and wayang, while using the city as a base for Borobudur, Prambanan, and Mount Merapi.",
            "What to Expect": "A compact, student-filled cultural city with a busy Malioboro corridor, palace traditions, inexpensive food, tropical heat, and frequent rain in the wet season.",
            "Recommended Stay": "3-4 days",
            "Avg. Cost/Day (3* Hotel & Food)": 45.0,
            "Rent a Car?": "No (hire a car and driver for temples and rural areas)",
            "Highlights": "Tour the Kraton and Taman Sari, shop for batik on Malioboro, explore Kotagede silver workshops, visit Prambanan at sunset, and make a sunrise trip to Borobudur or Mount Merapi.",
            "Introduction Sentence": "Yogyakarta is Java's royal and artistic heart, where palace culture, batik, street food, and world-class Hindu-Buddhist monuments meet a lively student city.",
        },
    },
    "Bishkek": {
        "selection": "no",
        "ideal": "May-October",
        "avoid": "December-February (cold and snowy)",
        "population": 1_200_000,
        "reachable": "Manas International Airport (FRU)",
        "visa": "no",
        "visa_requirement": "Visa-free for up to 60 days",
        "flight": 8.0,
        "profile": {
            "Country": "Kyrgyzstan",
            "Why to Go There": "See a leafy Central Asian capital with Silk Road connections, lively bazaars, Soviet-era avenues, and immediate access to Tian Shan mountain excursions.",
            "What to Expect": "A walkable, tree-lined city at about 800 metres with Russian and Kyrgyz influences, affordable local food, cold winters, hot summers, and mountains visible to the south.",
            "Recommended Stay": "2-3 days",
            "Avg. Cost/Day (3* Hotel & Food)": 40.0,
            "Rent a Car?": "Sometimes (for Ala-Archa and mountain day trips)",
            "Highlights": "Visit Ala-Too Square and the National Historical Museum, browse Osh Bazaar, walk Oak Park and Erkindik Boulevard, and take a day trip to Ala-Archa National Park or Burana Tower.",
            "Introduction Sentence": "Bishkek is Kyrgyzstan's green, mountain-framed capital, a relaxed base for bazaars, Soviet urban history, nomadic culture, and Tian Shan day trips.",
        },
    },
    "Dubai": {
        "selection": "no",
        "ideal": "November-March",
        "avoid": "June-August (extreme heat)",
        "population": 3_600_000,
        "reachable": "Dubai International Airport (DXB)",
        "visa": "no",
        "visa_requirement": "Visa-free for up to 90 days",
        "flight": 6.5,
        "profile": {
            "Why to Go There": "See the contrast between creekside trading districts and a globally recognised skyline, with major shopping, architecture, beaches, desert excursions, and museums.",
            "What to Expect": "A highly connected, multicultural desert metropolis with excellent metro and airport infrastructure, expensive attractions, strict public conduct rules, and extreme summer heat.",
            "Recommended Stay": "3-5 days",
            "Avg. Cost/Day (3* Hotel & Food)": 120.0,
            "Rent a Car?": "No (Metro and taxis cover the main visitor areas)",
            "Highlights": "Go up the Burj Khalifa, cross Dubai Creek by abra, explore Al Fahidi and Deira's Gold and Spice Souks, visit the Museum of the Future, and book a desert conservation or dune excursion.",
            "Introduction Sentence": "Dubai is a Gulf trading city transformed into a global hub, balancing creekside heritage and souks with the Burj Khalifa, luxury districts, beaches, and desert landscapes.",
        },
    },
    "Montevideo": {
        "selection": "no",
        "ideal": "October-April",
        "avoid": "June-August (coolest and windiest)",
        "population": 2_000_000,
        "reachable": "Carrasco International Airport (MVD)",
        "visa": "no",
        "visa_requirement": "Visa-free for up to 90 days",
        "flight": 14.0,
        "profile": {
            "Why to Go There": "Combine Río de la Plata waterfront life with colonial streets, Art Deco architecture, tango and candombe, football history, and Uruguay's relaxed café and grill culture.",
            "What to Expect": "A temperate, walkable capital with a long coastal Rambla, beaches within the city, a lively Ciudad Vieja, regular year-round rain, and cool windy winters.",
            "Recommended Stay": "3-4 days",
            "Avg. Cost/Day (3* Hotel & Food)": 80.0,
            "Rent a Car?": "No (buses, taxis, and walking cover the city)",
            "Highlights": "Walk Ciudad Vieja and Plaza Independencia, tour the Solís Theatre, eat at Mercado del Puerto, follow the Rambla through Pocitos, and visit the Estadio Centenario and Parque Rodó.",
            "Introduction Sentence": "Montevideo is Uruguay's laid-back Río de la Plata capital, known for its historic old town, Art Deco streets, beaches, long Rambla, tango, candombe, and excellent beef.",
        },
    },
}

CLIMATE_DATA = {
    "Panama City": {
        "high": [32.2, 32.7, 33.2, 33.4, 32.4, 31.8, 31.8, 31.9, 31.4, 31.1, 31.1, 31.7],
        "low": [21.4, 21.5, 21.8, 22.6, 22.9, 22.7, 22.6, 22.5, 22.4, 22.2, 22.1, 21.8],
        "rainy_days": [2.4, 1.6, 1.9, 5.2, 15.4, 16.8, 15.1, 16.0, 17.4, 19.5, 16.8, 7.5],
        "rain": [24.1, 12.3, 14.2, 71.0, 221.8, 242.0, 189.5, 221.0, 268.4, 311.2, 258.9, 124.6],
    },
    "San José": {
        "high": [28.2, 29.1, 29.9, 30.3, 28.8, 28.2, 28.2, 28.3, 27.8, 27.1, 27.2, 27.9],
        "low": [18.5, 18.7, 18.8, 19.1, 19.2, 19.0, 19.0, 18.8, 18.3, 18.5, 18.3, 18.3],
        "rainy_days": [3, 3, 5, 10, 23, 22, 20, 22, 26, 25, 17, 8],
        "rain": [6.3, 10.2, 13.8, 79.9, 267.6, 280.1, 181.5, 276.9, 355.1, 330.6, 135.5, 33.5],
    },
    "Yogyakarta": {
        "high": [29.8, 30.5, 31.3, 31.5, 31.1, 31.0, 30.3, 30.7, 31.5, 31.6, 30.9, 30.1],
        "low": [22.9, 22.8, 22.9, 23.0, 22.7, 21.5, 20.6, 20.6, 21.7, 22.7, 23.0, 22.8],
        "rainy_days": [18, 16, 15, 12, 8, 6, 5, 4, 5, 10, 15, 18],
        "rain": [392, 299, 363, 149, 141, 68, 29, 16, 49, 136, 237, 278],
    },
    "Bishkek": {
        "high": [2.9, 5.1, 12.1, 18.7, 24.1, 29.5, 32.4, 31.4, 25.6, 18.5, 10.3, 4.6],
        "low": [-7.1, -4.9, 1.0, 6.9, 11.2, 16.1, 18.4, 16.9, 11.7, 5.6, -0.5, -5.2],
        "rainy_days": [3, 5, 9, 12, 13, 10, 10, 6, 6, 8, 7, 4],
        "rain": [28, 36, 48, 71, 59, 34, 19, 15, 18, 37, 45, 37],
    },
    "Dubai": {
        "high": [24.0, 25.0, 30.0, 34.0, 37.5, 39.9, 41.7, 42.1, 39.5, 36.5, 31.0, 26.0],
        "low": [14.3, 15.5, 18.3, 21.7, 25.1, 26.9, 30.0, 30.4, 27.7, 24.1, 20.1, 16.3],
        "rainy_days": [5.5, 4.7, 5.8, 2.6, 0.3, 0.2, 0.5, 0.5, 0.1, 0.2, 1.3, 3.8],
        "rain": [18.8, 25.0, 22.1, 7.2, 0.4, 0.2, 0.8, 0.2, 0.0, 1.1, 2.7, 16.2],
    },
    "Montevideo": {
        "high": [27.8, 27.0, 25.3, 22.0, 18.5, 15.6, 14.7, 16.7, 17.9, 20.7, 23.7, 26.4],
        "low": [18.8, 18.6, 17.1, 14.1, 11.0, 8.1, 7.3, 8.5, 9.9, 12.4, 14.7, 17.1],
        "rainy_days": [6, 6, 6, 7, 6, 7, 6, 7, 7, 7, 7, 7],
        "rain": [94.6, 93.8, 105.8, 111.1, 83.4, 89.4, 93.2, 89.9, 92.1, 102.2, 95.9, 91.3],
    },
}


def populate_details() -> None:
    workbook = openpyxl.load_workbook(WORKBOOK)
    worksheet = workbook.active
    headers = {
        str(cell.value).strip(): cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }
    destination_column = headers["Destination"]
    monthly_headers = {
        "high": [headers[f"{month} High (C)"] for month in ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]],
        "low": [headers[f"{month} Low (C)"] for month in ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]],
        "rainy_days": [headers[f"{month} Rainy Days"] for month in ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]],
        "rain": [headers[f"{month} Rain (mm)"] for month in ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]],
        "aqi": [headers[f"{month} AQI"] for month in ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]],
    }

    fields = {
        "selection": "In näherer Auswahl 2025?",
        "ideal": "Ideal Time to Go",
        "avoid": "Avoid Going There",
        "population": "Population (Metro Area)",
        "reachable": "Reachable via",
        "visa": "Visa?",
        "visa_requirement": "Visa Requirement",
        "flight": "Flight Time to Frankfurt (hours)",
    }
    annual_fields = {
        "high": "Avg High Temp (°C)",
        "low": "Avg Low Temp (°C)",
        "rainy_days": "Avg Rainy Days/Month",
        "rain": "Avg Rain (mm/Month)",
        "aqi": "Avg AQI",
    }
    profile_fields = {
        "Why to Go There", "What to Expect", "Recommended Stay",
        "Avg. Cost/Day (3* Hotel & Food)", "Rent a Car?", "Highlights",
        "Introduction Sentence",
    }
    month_names = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    month_ratings = {
        "Panama City": ["ideal", "ideal", "ideal", "ideal", "ok", "bad", "bad", "bad", "bad", "bad", "bad", "ideal"],
        "San José": ["ideal", "ideal", "ideal", "ideal", "ok", "ok", "ok", "ok", "bad", "bad", "ok", "ideal"],
        "Yogyakarta": ["bad", "bad", "bad", "ok", "ok", "ideal", "ideal", "ideal", "ideal", "ok", "bad", "bad"],
        "Bishkek": ["bad", "bad", "ok", "ideal", "ideal", "ideal", "ideal", "ideal", "ideal", "ideal", "ok", "bad"],
        "Dubai": ["ideal", "ideal", "ideal", "ok", "bad", "bad", "bad", "bad", "bad", "ok", "ideal", "ideal"],
        "Montevideo": ["ideal", "ideal", "ideal", "ideal", "ok", "bad", "bad", "bad", "ok", "ideal", "ideal", "ideal"],
    }

    for row in range(2, worksheet.max_row + 1):
        destination = worksheet.cell(row, destination_column).value
        if destination not in DESTINATION_DETAILS:
            continue
        details = DESTINATION_DETAILS[destination]
        for key, header in fields.items():
            worksheet.cell(row, headers[header], details[key])
        for header, value in details["profile"].items():
            worksheet.cell(row, headers[header], value)
        climate = CLIMATE_DATA[destination]
        for key, columns in monthly_headers.items():
            if key == "aqi":
                continue
            for column, value in zip(columns, climate[key]):
                worksheet.cell(row, column, value)
        for key, header in annual_fields.items():
            values = climate[key] if key != "aqi" else [worksheet.cell(row, column).value for column in monthly_headers[key]]
            worksheet.cell(row, headers[header], round(sum(values) / len(values), 1))
        for month_name, climate_label in zip(month_names, month_ratings[destination]):
            worksheet.cell(row, headers[month_name], climate_label)
        print(f"Populated non-review fields: row {row} ({destination})")

    workbook.save(WORKBOOK)
    workbook.close()


if __name__ == "__main__":
    populate_details()
