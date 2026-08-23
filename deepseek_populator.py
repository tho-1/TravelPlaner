"""Write a DeepSeek-generated destination profile into the workbook.

- Computes the standardized 0-10 review score using the ``review_analyzer``
  methodology (aspect sentiment + Bayesian platform average).
- Appends a new row (destination / country / continent) and fills every profile
  column (details, month ratings, reviews, climate) in a single save.
- Raises ``WorkbookLockedError`` (from ``data_utils``) when the file is locked
  by another program, mirroring the add-destination flow.
"""

from __future__ import annotations

from typing import Tuple

import openpyxl

from data_utils import (
    DATA_PATH,
    WorkbookLockedError,
    _find_destination_sheet,
    load_destinations,
    update_reviews,
)
from deepseek_client import generate_destination_profile, generate_review_profile
from review_analyzer import ASPECTS, AspectSentiment, DestinationReviewData, PlatformRating

MONTHS_FULL = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]
MONTHS_SHORT = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def add_destination_with_deepseek(
    destination: str, country: str, continent: str, path=DATA_PATH
) -> Tuple[bool, str]:
    """Call the DeepSeek API and populate a brand-new workbook row.

    Returns ``(success, message)``. Raises ``WorkbookLockedError`` if the file
    is locked, and ``RuntimeError`` if the API call itself fails.
    """
    profile = generate_destination_profile(destination, country, continent)
    return populate_destination_with_ai(destination, country, continent, profile, path=path)


def populate_destination_with_ai(
    destination: str,
    country: str,
    continent: str,
    profile: dict,
    path=DATA_PATH,
) -> Tuple[bool, str]:
    """Append a new destination row and fill it from a generated profile dict.

    The profile is expected in the schema produced by
    ``deepseek_client.generate_destination_profile``. Missing/odd fields are
    skipped gracefully; a malformed review section never blocks the whole add.
    """
    dest_clean = str(destination).strip()
    if not dest_clean:
        return False, "Destination name cannot be empty."

    sheet_name = _find_destination_sheet(path)
    if sheet_name is None:
        return False, "Could not find destination sheet in workbook."

    try:
        wb = openpyxl.load_workbook(path)
    except PermissionError as exc:
        raise WorkbookLockedError(
            "Destinations.xlsx is currently open in another program (e.g. Excel). "
            "Please close the file and press Retry."
        ) from exc
    ws = wb[sheet_name]

    # Header map: exact header text -> 1-based column index.
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val is not None:
            headers[str(val).strip()] = col_idx

    dest_col = headers.get("Destination")
    if dest_col is None:
        wb.close()
        return False, "Could not find the 'Destination' column in the workbook."

    # Reject duplicates.
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=dest_col).value
        if val is not None and str(val).strip().lower() == dest_clean.lower():
            wb.close()
            return False, f"Destination '{dest_clean}' already exists in the workbook (Row {row_idx})."

    new_row = ws.max_row + 1
    review_score = _apply_profile_to_row(ws, headers, new_row, profile, country, continent, dest_clean)

    # Data Status intentionally left blank for AI-populated rows (not a placeholder).

    try:
        wb.save(path)
    except PermissionError as exc:
        wb.close()
        raise WorkbookLockedError(
            "Destinations.xlsx is currently open in another program (e.g. Excel). "
            "Please close the file and press Retry."
        ) from exc
    wb.close()

    from data_utils import _clear_destination_cache
    _clear_destination_cache()
    return True, (
        f"Destination '{dest_clean}' added and populated with DeepSeek AI "
        f"(review score {review_score}/10)."
    )


def _apply_profile_to_row(ws, headers, row_idx, profile, country, continent, dest_clean):
    """Write every generated profile field onto ``row_idx``.

    Shared by the add-new flow and the populate-existing flow. Returns the
    computed review score (0-10) or ``None`` if the review section was missing.
    """
    def _ensure_header(name: str) -> int:
        if name not in headers:
            col = ws.max_column + 1
            ws.cell(row=1, column=col, value=name)
            headers[name] = col
        return headers[name]

    def _set(name: str, value) -> None:
        if value is None:
            return
        if isinstance(value, str) and not value.strip():
            return
        ws.cell(row=row_idx, column=_ensure_header(name), value=value)

    def _num(value):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    # ── Identity fields ────────────────────────────────────────────────────────
    _set("Destination", dest_clean)
    _set("Continent", profile.get("continent") or continent)
    _set("Country", profile.get("country") or country)

    prof = profile.get("profile") or {}

    # ── Written profile fields ─────────────────────────────────────────────────
    _set("Why to Go There", prof.get("why_to_go"))
    _set("What to Expect", prof.get("what_to_expect"))
    _set("Ideal Time to Go", prof.get("ideal_time_to_go"))
    _set("Avoid Going There", prof.get("avoid_going"))
    _set("Recommended Stay", prof.get("recommended_stay"))
    _set("Avg. Cost/Day (3* Hotel & Food)", _num(prof.get("avg_cost_per_day_eur")))
    _set("Rent a Car?", prof.get("rent_a_car"))
    _set("Highlights", prof.get("highlights"))
    _set("Population (Metro Area)", _num(prof.get("population_metro")))
    # NOTE: "Reachable via" is intentionally NOT populated anymore (column
    # removed from the workbook / no longer needed).
    _set("Visa?", prof.get("visa_requirement"))
    _set("Visa Requirement", prof.get("visa_requirement"))
    _set("Flight Time to Frankfurt (hours)", _num(prof.get("flight_time_hours")))
    _set("Safety Rating (10 = safest)", _num(prof.get("safety_rating_10")))
    _set("Introduction Sentence", prof.get("introduction_sentence"))

    in_eu = prof.get("in_eu")
    if isinstance(in_eu, str):
        in_eu = in_eu.strip().lower() in {"true", "yes", "1", "y"}
    if isinstance(in_eu, bool):
        _set("Yes?", in_eu)

    # ── Month ratings (January..December -> ideal/ok/bad) ──────────────────────
    month_ratings = prof.get("month_ratings") or {}
    month_lookup = {str(k).strip().lower(): v for k, v in month_ratings.items()}
    for month in MONTHS_FULL:
        val = month_lookup.get(month.lower())
        if isinstance(val, str) and val.strip().lower() in {"ideal", "ok", "bad"}:
            _set(month, val.strip().lower())

    # ── Reviews (review_analyzer methodology) ──────────────────────────────────
    review_score = None
    try:
        rev = profile.get("reviews") or {}
        aspect_counts = {}
        aspects_raw = rev.get("aspects") or {}
        for aspect in ASPECTS:
            counts = aspects_raw.get(aspect) or {}
            aspect_counts[aspect] = AspectSentiment(
                pos=int(counts.get("pos", 0) or 0),
                neutral=int(counts.get("neutral", 0) or 0),
                neg=int(counts.get("neg", 0) or 0),
            )

        platform_ratings = []
        for rating in rev.get("platform_ratings") or []:
            platform_ratings.append(
                PlatformRating(
                    source_name=str(rating.get("source_name", "Unknown")),
                    avg_rating=float(rating.get("avg_rating", 0.0) or 0.0),
                    review_count=int(rating.get("review_count", 0) or 0),
                )
            )

        total = int(rev.get("total_reviews_analyzed", 50) or 50)
        date_range = str(rev.get("date_range", "Feb 2024 - Feb 2026"))

        review_data = DestinationReviewData(
            destination_name=dest_clean,
            country=profile.get("country") or country,
            aspect_counts=aspect_counts,
            platform_ratings=platform_ratings,
            total_reviews_analyzed=total,
            date_range=date_range,
        )
        result = review_data.calculate_final_scores()
        review_score = result["Review Score (0-10)"]

        praise = str(rev.get("praise", "") or "").strip()
        dislikes = str(rev.get("dislikes", "") or "").strip()
        tourist_summary = (
            f"{praise} Recurring criticisms mention {dislikes} "
            f"This assessment is based on {total} traveler reviews from {date_range}, "
            f"with {result['Confidence']} confidence."
        )

        _set("Reviews", review_score)
        _set("Tourist Reviews", tourist_summary)
        _set("What do the reviews praise?", praise)
        _set("What do they dislike?", dislikes)
    except Exception:
        # Never let review-score quirks block the add; the written profile and
        # climate data are still valuable.
        pass

    # ── Monthly climate (Jan..Dec High/Low/Rainy Days/Rain/AQI) ───────────────
    climate = profile.get("climate") or {}
    climate_series = {
        "High (C)": climate.get("high_c"),
        "Low (C)": climate.get("low_c"),
        "Rainy Days": climate.get("rainy_days"),
        "Rain (mm)": climate.get("rain_mm"),
        "AQI": climate.get("aqi"),
    }
    for suffix, values in climate_series.items():
        if not isinstance(values, (list, tuple)) or len(values) < 12:
            continue
        for i, month in enumerate(MONTHS_SHORT):
            _set(f"{month} {suffix}", _num(values[i]))

    # ── Annual climate average (Avg AQI only) ─────────────────────────────────
    # The other yearly-average columns (Avg High Temp, Avg Low Temp, Avg Rainy
    # Days, Avg Rain mm) were removed from the workbook and are NOT populated.
    def _avg(values):
        nums = [_num(v) for v in values] if isinstance(values, (list, tuple)) else []
        nums = [n for n in nums if n is not None]
        return round(sum(nums) / len(nums), 1) if nums else None

    _set("Avg AQI", _avg(climate.get("aqi")))

    return review_score


def populate_existing_destination_with_ai(
    destination: str, country: str, continent: str, path=DATA_PATH
) -> Tuple[bool, str]:
    """Populate an EXISTING destination row (e.g. a placeholder) via DeepSeek.

    Returns ``(success, message)``. Raises ``WorkbookLockedError`` if the file
    is locked, and ``RuntimeError`` if the API call itself fails.
    """
    dest_clean = str(destination).strip()
    if not dest_clean:
        return False, "Destination name cannot be empty."

    profile = generate_destination_profile(destination, country, continent)
    return populate_existing_destination_with_profile(dest_clean, country, continent, profile, path=path)


def populate_existing_destination_with_profile(
    destination: str,
    country: str,
    continent: str,
    profile: dict,
    path=DATA_PATH,
) -> Tuple[bool, str]:
    """Fill an existing destination row from a generated profile dict.

    Finds the row by destination name, writes all profile fields (details,
    month ratings, reviews, climate), clears the placeholder ``Data Status``
    so the "incomplete" warning disappears, and saves once.
    """
    dest_clean = str(destination).strip()
    if not dest_clean:
        return False, "Destination name cannot be empty."

    sheet_name = _find_destination_sheet(path)
    if sheet_name is None:
        return False, "Could not find destination sheet in workbook."

    try:
        wb = openpyxl.load_workbook(path)
    except PermissionError as exc:
        raise WorkbookLockedError(
            "Destinations.xlsx is currently open in another program (e.g. Excel). "
            "Please close the file and press Retry."
        ) from exc
    ws = wb[sheet_name]

    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val is not None:
            headers[str(val).strip()] = col_idx

    dest_col = headers.get("Destination")
    if dest_col is None:
        wb.close()
        return False, "Could not find the 'Destination' column in the workbook."

    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=dest_col).value
        if val is not None and str(val).strip().lower() == dest_clean.lower():
            target_row = row_idx
            break

    if target_row is None:
        wb.close()
        return False, f"Destination '{dest_clean}' was not found in the workbook."

    review_score = _apply_profile_to_row(ws, headers, target_row, profile, country, continent, dest_clean)

    # Clear the placeholder status so the "incomplete" warning disappears.
    # NOTE: use `.value = None` (not `cell(..., value=None)`) — openpyxl treats
    # a `None` value kwarg as "leave unchanged", so it would NOT clear the cell.
    status_col = headers.get("Data Status")
    if status_col:
        ws.cell(row=target_row, column=status_col).value = None

    try:
        wb.save(path)
    except PermissionError as exc:
        wb.close()
        raise WorkbookLockedError(
            "Destinations.xlsx is currently open in another program (e.g. Excel). "
            "Please close the file and press Retry."
        ) from exc
    wb.close()

    from data_utils import _clear_destination_cache
    _clear_destination_cache()
    return True, (
        f"Destination '{dest_clean}' populated with DeepSeek AI "
        f"(review score {review_score}/10)."
    )


def refresh_reviews_with_deepseek(
    destination: str, country: str, path=DATA_PATH
) -> Tuple[bool, str]:
    """Refresh only review-related fields for an existing destination."""
    review = generate_review_profile(destination, country)
    aspect_counts = {}
    for aspect in ASPECTS:
        counts = review.get("aspects", {}).get(aspect, {})
        aspect_counts[aspect] = AspectSentiment(
            pos=int(counts.get("pos", 0)),
            neutral=int(counts.get("neutral", 0)),
            neg=int(counts.get("neg", 0)),
        )
    platform_ratings = [
        PlatformRating(
            source_name=str(item.get("source_name", "Unknown")),
            avg_rating=float(item.get("avg_rating", 0)),
            review_count=int(item.get("review_count", 0)),
        )
        for item in review.get("platform_ratings", [])
    ]
    review_data = DestinationReviewData(
        destination_name=destination,
        country=country,
        aspect_counts=aspect_counts,
        platform_ratings=platform_ratings,
        total_reviews_analyzed=int(review.get("total_reviews_analyzed", 50)),
        date_range=str(review.get("date_range", "Unknown")),
    )
    result = review_data.calculate_final_scores()
    praise = str(review["praise"]).strip()
    dislikes = str(review["dislikes"]).strip()
    summary = (
        f"{praise} Recurring criticisms mention {dislikes} "
        f"This assessment is based on {review_data.total_reviews_analyzed} traveler reviews "
        f"from {review_data.date_range}, with {result['Confidence']} confidence."
    )
    update_reviews(
        destination,
        result["Review Score (0-10)"],
        summary,
        praise,
        dislikes,
        path=path,
    )
    return True, f"{destination}: review score {result['Review Score (0-10)']}/10"
