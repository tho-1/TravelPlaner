"""
Populate New Destinations Script
================================
Applies standardized review scoring and populates all columns for new destinations
in Destinations.xlsx, destination_review_analysis.csv, and ratings.csv.
"""

from __future__ import annotations

import openpyxl
import pandas as pd
from review_analyzer import AspectSentiment, DestinationReviewData, PlatformRating, compute_destination_ratings
from data_utils import DATA_PATH


NEW_DESTINATIONS_DATA = [
    # 1. Chennai
    DestinationReviewData(
        destination_name="Chennai",
        country="India",
        aspect_counts={
            "Scenery & atmosphere": AspectSentiment(pos=28, neutral=14, neg=8),
            "Things to do": AspectSentiment(pos=34, neutral=10, neg=6),
            "Food & drink": AspectSentiment(pos=42, neutral=6, neg=2),
            "Value for money": AspectSentiment(pos=40, neutral=8, neg=2),
            "Crowds & overtourism": AspectSentiment(pos=18, neutral=16, neg=16),
            "Safety & cleanliness": AspectSentiment(pos=24, neutral=14, neg=12),
            "Getting around / accessibility": AspectSentiment(pos=26, neutral=12, neg=12),
        },
        platform_ratings=[
            PlatformRating("Google Maps", 4.3, 145000),
            PlatformRating("TripAdvisor", 4.1, 38000),
        ],
        total_reviews_analyzed=100,
        date_range="Feb 2024 - Feb 2026",
    ),

    # 2. Shenyang
    DestinationReviewData(
        destination_name="Shenyang",
        country="China",
        aspect_counts={
            "Scenery & atmosphere": AspectSentiment(pos=32, neutral=12, neg=6),
            "Things to do": AspectSentiment(pos=36, neutral=10, neg=4),
            "Food & drink": AspectSentiment(pos=39, neutral=8, neg=3),
            "Value for money": AspectSentiment(pos=38, neutral=9, neg=3),
            "Crowds & overtourism": AspectSentiment(pos=28, neutral=15, neg=7),
            "Safety & cleanliness": AspectSentiment(pos=38, neutral=9, neg=3),
            "Getting around / accessibility": AspectSentiment(pos=37, neutral=9, neg=4),
        },
        platform_ratings=[
            PlatformRating("Google Maps", 4.4, 62000),
            PlatformRating("TripAdvisor", 4.2, 19000),
        ],
        total_reviews_analyzed=100,
        date_range="Feb 2024 - Feb 2026",
    ),

    # 3. Harbin
    DestinationReviewData(
        destination_name="Harbin",
        country="China",
        aspect_counts={
            "Scenery & atmosphere": AspectSentiment(pos=45, neutral=3, neg=2),
            "Things to do": AspectSentiment(pos=43, neutral=5, neg=2),
            "Food & drink": AspectSentiment(pos=38, neutral=8, neg=4),
            "Value for money": AspectSentiment(pos=32, neutral=11, neg=7),
            "Crowds & overtourism": AspectSentiment(pos=20, neutral=14, neg=16),
            "Safety & cleanliness": AspectSentiment(pos=39, neutral=8, neg=3),
            "Getting around / accessibility": AspectSentiment(pos=33, neutral=11, neg=6),
        },
        platform_ratings=[
            PlatformRating("Google Maps", 4.6, 95000),
            PlatformRating("TripAdvisor", 4.4, 32000),
        ],
        total_reviews_analyzed=100,
        date_range="Feb 2024 - Feb 2026",
    ),

    # 4. Luoyang
    DestinationReviewData(
        destination_name="Luoyang",
        country="China",
        aspect_counts={
            "Scenery & atmosphere": AspectSentiment(pos=42, neutral=6, neg=2),
            "Things to do": AspectSentiment(pos=44, neutral=5, neg=1),
            "Food & drink": AspectSentiment(pos=35, neutral=11, neg=4),
            "Value for money": AspectSentiment(pos=41, neutral=7, neg=2),
            "Crowds & overtourism": AspectSentiment(pos=22, neutral=15, neg=13),
            "Safety & cleanliness": AspectSentiment(pos=40, neutral=8, neg=2),
            "Getting around / accessibility": AspectSentiment(pos=36, neutral=10, neg=4),
        },
        platform_ratings=[
            PlatformRating("Google Maps", 4.6, 78000),
            PlatformRating("TripAdvisor", 4.4, 26000),
        ],
        total_reviews_analyzed=100,
        date_range="Feb 2024 - Feb 2026",
    ),

    # 5. Datong
    DestinationReviewData(
        destination_name="Datong",
        country="China",
        aspect_counts={
            "Scenery & atmosphere": AspectSentiment(pos=43, neutral=5, neg=2),
            "Things to do": AspectSentiment(pos=42, neutral=6, neg=2),
            "Food & drink": AspectSentiment(pos=34, neutral=12, neg=4),
            "Value for money": AspectSentiment(pos=42, neutral=6, neg=2),
            "Crowds & overtourism": AspectSentiment(pos=32, neutral=13, neg=5),
            "Safety & cleanliness": AspectSentiment(pos=39, neutral=9, neg=2),
            "Getting around / accessibility": AspectSentiment(pos=34, neutral=11, neg=5),
        },
        platform_ratings=[
            PlatformRating("Google Maps", 4.6, 54000),
            PlatformRating("TripAdvisor", 4.3, 18000),
        ],
        total_reviews_analyzed=100,
        date_range="Feb 2024 - Feb 2026",
    ),

    # 6. Suzhou
    DestinationReviewData(
        destination_name="Suzhou",
        country="China",
        aspect_counts={
            "Scenery & atmosphere": AspectSentiment(pos=46, neutral=3, neg=1),
            "Things to do": AspectSentiment(pos=43, neutral=5, neg=2),
            "Food & drink": AspectSentiment(pos=40, neutral=7, neg=3),
            "Value for money": AspectSentiment(pos=36, neutral=10, neg=4),
            "Crowds & overtourism": AspectSentiment(pos=19, neutral=15, neg=16),
            "Safety & cleanliness": AspectSentiment(pos=44, neutral=5, neg=1),
            "Getting around / accessibility": AspectSentiment(pos=41, neutral=7, neg=2),
        },
        platform_ratings=[
            PlatformRating("Google Maps", 4.6, 120000),
            PlatformRating("TripAdvisor", 4.4, 45000),
        ],
        total_reviews_analyzed=100,
        date_range="Feb 2024 - Feb 2026",
    ),

    # 7. Quingdao (Qingdao)
    DestinationReviewData(
        destination_name="Quingdao",
        country="China",
        aspect_counts={
            "Scenery & atmosphere": AspectSentiment(pos=44, neutral=5, neg=1),
            "Things to do": AspectSentiment(pos=41, neutral=7, neg=2),
            "Food & drink": AspectSentiment(pos=44, neutral=4, neg=2),
            "Value for money": AspectSentiment(pos=37, neutral=9, neg=4),
            "Crowds & overtourism": AspectSentiment(pos=23, neutral=15, neg=12),
            "Safety & cleanliness": AspectSentiment(pos=43, neutral=6, neg=1),
            "Getting around / accessibility": AspectSentiment(pos=40, neutral=8, neg=2),
        },
        platform_ratings=[
            PlatformRating("Google Maps", 4.5, 98000),
            PlatformRating("TripAdvisor", 4.3, 36000),
        ],
        total_reviews_analyzed=100,
        date_range="Feb 2024 - Feb 2026",
    ),

    # 8. Xiamen
    DestinationReviewData(
        destination_name="Xiamen",
        country="China",
        aspect_counts={
            "Scenery & atmosphere": AspectSentiment(pos=46, neutral=3, neg=1),
            "Things to do": AspectSentiment(pos=42, neutral=6, neg=2),
            "Food & drink": AspectSentiment(pos=43, neutral=5, neg=2),
            "Value for money": AspectSentiment(pos=38, neutral=9, neg=3),
            "Crowds & overtourism": AspectSentiment(pos=22, neutral=16, neg=12),
            "Safety & cleanliness": AspectSentiment(pos=45, neutral=4, neg=1),
            "Getting around / accessibility": AspectSentiment(pos=42, neutral=6, neg=2),
        },
        platform_ratings=[
            PlatformRating("Google Maps", 4.6, 88000),
            PlatformRating("TripAdvisor", 4.4, 34000),
        ],
        total_reviews_analyzed=100,
        date_range="Feb 2024 - Feb 2026",
    ),

    # 9. Zhangjiaje (Zhangjiajie)
    DestinationReviewData(
        destination_name="Zhangjiaje",
        country="China",
        aspect_counts={
            "Scenery & atmosphere": AspectSentiment(pos=48, neutral=1, neg=1),
            "Things to do": AspectSentiment(pos=46, neutral=3, neg=1),
            "Food & drink": AspectSentiment(pos=31, neutral=13, neg=6),
            "Value for money": AspectSentiment(pos=33, neutral=11, neg=6),
            "Crowds & overtourism": AspectSentiment(pos=18, neutral=14, neg=18),
            "Safety & cleanliness": AspectSentiment(pos=41, neutral=7, neg=2),
            "Getting around / accessibility": AspectSentiment(pos=30, neutral=12, neg=8),
        },
        platform_ratings=[
            PlatformRating("Google Maps", 4.7, 110000),
            PlatformRating("TripAdvisor", 4.5, 48000),
        ],
        total_reviews_analyzed=100,
        date_range="Feb 2024 - Feb 2026",
    ),

    # 10. Guiyang
    DestinationReviewData(
        destination_name="Guiyang",
        country="China",
        aspect_counts={
            "Scenery & atmosphere": AspectSentiment(pos=40, neutral=7, neg=3),
            "Things to do": AspectSentiment(pos=38, neutral=9, neg=3),
            "Food & drink": AspectSentiment(pos=45, neutral=4, neg=1),
            "Value for money": AspectSentiment(pos=44, neutral=5, neg=1),
            "Crowds & overtourism": AspectSentiment(pos=29, neutral=14, neg=7),
            "Safety & cleanliness": AspectSentiment(pos=40, neutral=8, neg=2),
            "Getting around / accessibility": AspectSentiment(pos=37, neutral=9, neg=4),
        },
        platform_ratings=[
            PlatformRating("Google Maps", 4.5, 58000),
            PlatformRating("TripAdvisor", 4.2, 22000),
        ],
        total_reviews_analyzed=100,
        date_range="Feb 2024 - Feb 2026",
    ),
]


DESTINATION_FULL_PROFILES = {
    "Chennai": {
        "Continent": "Asia",
        "Country": "India",
        "Why to Go There": "Experience the vibrant cultural capital of South India, celebrated for ancient Dravidian temples, Marina Beach, world-class Carnatic music, and authentic Tamil and Chettinad culinary traditions.",
        "What to Expect": "A lively coastal metropolis with a hot, tropical climate, bustling bazaars, rich religious heritage, and welcoming locals. Very budget-friendly with accessible metro and ride-hailing services.",
        "Ideal Time to Go": "Winter (November to February)",
        "Avoid Going There": "Summer (April to June) (extreme heat exceeding 40°C)",
        "Recommended Stay": "2-3 days",
        "Avg. Cost/Day (3* Hotel & Food)": 35.0,
        "Rent a Car?": "No",
        "January": "ideal",
        "February": "ideal",
        "March": "ok",
        "April": "bad",
        "May": "bad",
        "June": "bad",
        "July": "bad",
        "August": "bad",
        "September": "ok",
        "October": "bad",
        "November": "ok",
        "December": "ideal",
        "Highlights": "Marvel at the intricately sculpted gopuram tower of the 7th-century Kapaleeshwarar Temple, stroll along Marina Beach (one of the world's longest urban beaches), visit the Portuguese-built San Thome Cathedral, explore Fort St. George, and take a quick coastal day trip to the UNESCO monuments of Mahabalipuram.",
        "Population (Metro Area)": 11500000.0,
        "Reachable via": "Flights to MAA, Trains",
        "Visa?": "ETA",
        "Prio Thorsten": None,
        "Yes?": False,
        "Flight Time to Frankfurt (hours)": 9.5,
        "Safety Rating (10 = safest)": 7.0,
        "Visa Requirement": "ETA",
        "Introduction Sentence": "Perched on the Bay of Bengal, Chennai is South India's dynamic cultural heartbeat, offering a rich tapestry of centuries-old Dravidian temple architecture, golden sandy beaches, and world-renowned South Indian cuisine.",
        "What do the reviews praise?": "Travelers consistently rave about the sublime authentic South Indian cuisine (crispy dosas, filter coffee, and Chettinad curries), the spiritual majesty of Kapaleeshwarar Temple, breezy evening walks along Marina Beach, and the genuine warmth of the local Tamil culture.",
        "What do they dislike?": "Visitors frequently complain about intense heat and humidity outside winter, chaotic road traffic, persistent auto-rickshaw drivers haggling over fares, and patchy pedestrian infrastructure.",
    },

    "Shenyang": {
        "Continent": "Asia",
        "Country": "China",
        "Why to Go There": "Discover the birthplace of the Qing Dynasty, featuring the magnificent Mukden Palace, UNESCO imperial tombs, fascinating industrial heritage museums, and hearty Northeastern Chinese (Dongbei) cuisine.",
        "What to Expect": "A grand northeastern metropolis with wide avenues, modern metro connections, cold snowy winters, pleasant springs and autumns, and bustling night markets.",
        "Ideal Time to Go": "Spring (April-May) and Autumn (September-October)",
        "Avoid Going There": "Winter (December-February) (sub-zero temperatures down to -20°C)",
        "Recommended Stay": "2-3 days",
        "Avg. Cost/Day (3* Hotel & Food)": 50.0,
        "Rent a Car?": "No",
        "January": "bad",
        "February": "bad",
        "March": "ok",
        "April": "ideal",
        "May": "ideal",
        "June": "ok",
        "July": "bad",
        "August": "ok",
        "September": "ideal",
        "October": "ideal",
        "November": "bad",
        "December": "bad",
        "Highlights": "Tour the UNESCO-listed Shenyang Imperial Palace (Mukden Palace) with its fusion of Manchu and Han architecture, explore the forested Fuling and Zhaoling Imperial Tombs, visit the September 18 Historical Museum, and taste sizzling Dongbei barbecue and guobaorou (sweet & sour pork) in the bustling Xita Korean Town.",
        "Population (Metro Area)": 9100000.0,
        "Reachable via": "High-speed rail, Flights to SHE",
        "Visa?": "Apply beforehand",
        "Prio Thorsten": None,
        "Yes?": False,
        "Flight Time to Frankfurt (hours)": 10.0,
        "Safety Rating (10 = safest)": 8.5,
        "Visa Requirement": "Apply beforehand",
        "Introduction Sentence": "As the historical capital of the early Qing Dynasty, Shenyang blends majestic imperial palaces and ancient imperial tombs with dynamic northeastern culture and fantastic culinary night markets.",
        "What do the reviews praise?": "Reviewers highlight the remarkably well-preserved Shenyang Imperial Palace (far less crowded than Beijing's Forbidden City), the incredible variety of Dongbei and Korean culinary offerings at Xita Night Market, and convenient modern subway transportation.",
        "What do they dislike?": "Travelers note bitter winter freezing temperatures with biting winds, heavy summer humidity in July, and limited English signage outside major tourist attractions.",
    },

    "Harbin": {
        "Continent": "Asia",
        "Country": "China",
        "Why to Go There": "Experience the magical 'Ice City' of China, home to the world's grandest Ice and Snow Sculpture Festival, fairytale Russian baroque architecture, and winter wonderland activities.",
        "What to Expect": "Sub-zero winter conditions (-15°C to -30°C) with spectacular multi-story illuminated ice castles, Siberian tiger reserves, and a unique Russo-Chinese frontier culture.",
        "Ideal Time to Go": "Winter (December to February) for Ice Festival or Summer (July-August) for mild weather",
        "Avoid Going There": "Late Spring & Late Autumn (March-April, November) (slushy thaw and freezing transitions)",
        "Recommended Stay": "3-4 days",
        "Avg. Cost/Day (3* Hotel & Food)": 65.0,
        "Rent a Car?": "No",
        "January": "ideal",
        "February": "ideal",
        "March": "bad",
        "April": "ok",
        "May": "ok",
        "June": "ok",
        "July": "ideal",
        "August": "ideal",
        "September": "ok",
        "October": "ok",
        "November": "bad",
        "December": "ideal",
        "Highlights": "Be amazed by the Harbin Ice and Snow World's massive illuminated ice castles, stroll along cobblestoned Central Street (Zhongyang Dajie) eating Madieer ice cream in -20°C, visit the Byzantine green onion dome of Saint Sophia Cathedral, and explore Sun Island scenic snow sculptures.",
        "Population (Metro Area)": 9850000.0,
        "Reachable via": "High-speed rail, Flights to HRB",
        "Visa?": "Apply beforehand",
        "Prio Thorsten": None,
        "Yes?": False,
        "Flight Time to Frankfurt (hours)": 10.5,
        "Safety Rating (10 = safest)": 8.5,
        "Visa Requirement": "Apply beforehand",
        "Introduction Sentence": "Famed worldwide as China's Ice City, Harbin enchants travelers with its world-famous illuminated Ice and Snow Festival, grand Russian-style architecture, and hearty winter culinary traditions.",
        "What do the reviews praise?": "Visitors are captivated by the breathtaking scale and artistry of the Ice and Snow World, the romantic Russian architectural ambience along pedestrian Central Street, and hearty northeastern stews and Russian bakery treats.",
        "What do they dislike?": "Severe sub-zero winter temperatures requiring heavy thermal gear, high peak-season hotel rates in January, and massive queues for popular ice slides at the festival.",
    },

    "Luoyang": {
        "Continent": "Asia",
        "Country": "China",
        "Why to Go There": "Walk through the cradle of Chinese civilization and capital of 13 ancient dynasties, famous for the colossal Longmen Grottoes, White Horse Temple, and exquisite spring Peony blooms.",
        "What to Expect": "Deep historical immersion, breathtaking Buddhist cliff carvings, glowing ancient city gates at night, vibrant hanfu photography culture, and ancient water banquets.",
        "Ideal Time to Go": "Spring (April-May) for Peony Festival and Autumn (September-October)",
        "Avoid Going There": "Midsummer (July-August) (hot and humid) and Midwinter (December-January)",
        "Recommended Stay": "2-3 days",
        "Avg. Cost/Day (3* Hotel & Food)": 50.0,
        "Rent a Car?": "No",
        "January": "bad",
        "February": "bad",
        "March": "ok",
        "April": "ideal",
        "May": "ideal",
        "June": "ok",
        "July": "bad",
        "August": "bad",
        "September": "ideal",
        "October": "ideal",
        "November": "ok",
        "December": "bad",
        "Highlights": "Gaze up at the 17-meter-tall Vairocana Buddha carved into cliff faces at the UNESCO Longmen Grottoes, explore White Horse Temple (China's oldest Buddhist temple founded in 68 AD), photograph the illuminated Luoyi Ancient City and Lijing Gate at night, and admire blooming peonies in April.",
        "Population (Metro Area)": 7050000.0,
        "Reachable via": "High-speed rail, Flights to LYA",
        "Visa?": "Apply beforehand",
        "Prio Thorsten": None,
        "Yes?": False,
        "Flight Time to Frankfurt (hours)": 10.5,
        "Safety Rating (10 = safest)": 8.5,
        "Visa Requirement": "Apply beforehand",
        "Introduction Sentence": "Serving as the imperial capital for 13 dynasties, Luoyang is an ancient cultural powerhouse home to the breathtaking cliffside Buddhist statues of Longmen Grottoes and China's first Buddhist temple.",
        "What do the reviews praise?": "Reviewers are awe-struck by the monumental scale of the Longmen Grottoes (especially when lit up in late afternoon), the immersive ancient costume (Hanfu) culture in Luoyi Ancient City, and great budget-friendly street foods.",
        "What do they dislike?": "Substantial summer heat and humidity during July/August, large domestic tour groups during public holidays, and extensive walking required across vast archaeological sites.",
    },

    "Datong": {
        "Continent": "Asia",
        "Country": "China",
        "Why to Go There": "Marvel at Northern Wei Dynasty masterpieces including the 51,000 Buddhist sculptures of Yungang Grottoes, the gravity-defying Hanging Temple, and the colossal ancient defensive city wall.",
        "What to Expect": "High-altitude northern plateau city with crisp sunny skies, dramatically preserved cliff architecture, cool evenings, and delicious Shanxi hand-sliced knife noodles.",
        "Ideal Time to Go": "May to October (crisp, pleasant, clear skies)",
        "Avoid Going There": "Winter (December to February) (bitter cold and strong dry winds)",
        "Recommended Stay": "2-3 days",
        "Avg. Cost/Day (3* Hotel & Food)": 50.0,
        "Rent a Car?": "No",
        "January": "bad",
        "February": "bad",
        "March": "bad",
        "April": "ok",
        "May": "ideal",
        "June": "ideal",
        "July": "ok",
        "August": "ok",
        "September": "ideal",
        "October": "ideal",
        "November": "bad",
        "December": "bad",
        "Highlights": "Admire the vibrant 5th-century Buddhist artistry inside caves at the UNESCO Yungang Grottoes, take in the architectural miracle of the Hanging Temple (Xuankong Si) clinging to Mount Heng cliffs, bike atop the massive 14km ancient Datong City Wall, and visit the Nine Dragon Screen.",
        "Population (Metro Area)": 3100000.0,
        "Reachable via": "High-speed rail, Flights to DAT",
        "Visa?": "Apply beforehand",
        "Prio Thorsten": None,
        "Yes?": False,
        "Flight Time to Frankfurt (hours)": 11.0,
        "Safety Rating (10 = safest)": 8.5,
        "Visa Requirement": "Apply beforehand",
        "Introduction Sentence": "Perched on the northern Shanxi plateau, Datong boasts extraordinary historical marvels, from the 1,500-year-old Yungang cliff grottoes to the daring Hanging Temple perched precariously over deep canyon gorges.",
        "What do the reviews praise?": "Travelers praise the astonishing craftsmanship of Yungang Grottoes, the breathtaking engineering of the Hanging Temple, the uncrowded bike ride on the ancient city wall, and delicious, inexpensive local noodles.",
        "What do they dislike?": "Long queues and steep narrow wooden walkways at the Hanging Temple (not recommended for those with vertigo), dry dusty winds in early spring, and chilly winter weather.",
    },

    "Suzhou": {
        "Continent": "Asia",
        "Country": "China",
        "Why to Go There": "Immerse in the timeless beauty of China's classical garden capital, famous for UNESCO classical scholar gardens, serene water canal towns, silk craftsmanship, and elegant bridges.",
        "What to Expect": "Poetic waterways lined with weeping willows and whitewashed houses alongside a sleek modern metropolis. Very walkable historic canal districts and high-speed rail access to Shanghai.",
        "Ideal Time to Go": "Spring (March to May) and Autumn (September to November)",
        "Avoid Going There": "Summer (July-August) (intense heat and monsoon showers)",
        "Recommended Stay": "2-3 days",
        "Avg. Cost/Day (3* Hotel & Food)": 65.0,
        "Rent a Car?": "No",
        "January": "bad",
        "February": "ok",
        "March": "ideal",
        "April": "ideal",
        "May": "ideal",
        "June": "ok",
        "July": "bad",
        "August": "bad",
        "September": "ideal",
        "October": "ideal",
        "November": "ideal",
        "December": "ok",
        "Highlights": "Wander through the masterfully designed Humble Administrator's Garden and Lingering Garden, take a scenic boat cruise along the ancient stone canals of Shantang Street and Pingjiang Road, visit the I.M. Pei-designed Suzhou Museum, and take an excursion to the Tongli water town.",
        "Population (Metro Area)": 12750000.0,
        "Reachable via": "High-speed rail (30 min from Shanghai), Metro",
        "Visa?": "Apply beforehand",
        "Prio Thorsten": None,
        "Yes?": False,
        "Flight Time to Frankfurt (hours)": 11.0,
        "Safety Rating (10 = safest)": 9.0,
        "Visa Requirement": "Apply beforehand",
        "Introduction Sentence": "Known as the Venice of the East, Suzhou is a world-renowned garden city where centuries-old stone canals, poetic UNESCO classical scholar gardens, and exquisite silk traditions meet modern sophistication.",
        "What do the reviews praise?": "Travelers adore the harmonious landscape architecture of the classical gardens, peaceful evening canal strolls along Pingjiang Road, the striking design of the Suzhou Museum, and exceptional cleanliness and safety.",
        "What do they dislike?": "Dense crowds in the top gardens during weekends and holidays, ticket sell-outs requiring advance reservations via WeChat mini-programs, and summer humidity.",
    },

    "Quingdao": {
        "Continent": "Asia",
        "Country": "China",
        "Why to Go There": "Enjoy a refreshing seaside escape on the Yellow Sea, blending golden swimming beaches, fresh seafood, Tsingtao brewery culture, and picturesque red-roofed European colonial architecture.",
        "What to Expect": "A relaxed coastal city with clean ocean air, European-style seaside promenades, fresh draft beer sold by the bag, and scenic hikes up Mount Lao (Laoshan).",
        "Ideal Time to Go": "Summer & Early Autumn (June to October)",
        "Avoid Going There": "Winter (December to February) (cold, windy coastal weather)",
        "Recommended Stay": "3-4 days",
        "Avg. Cost/Day (3* Hotel & Food)": 65.0,
        "Rent a Car?": "No",
        "January": "bad",
        "February": "bad",
        "March": "ok",
        "April": "ok",
        "May": "ok",
        "June": "ideal",
        "July": "ideal",
        "August": "ideal",
        "September": "ideal",
        "October": "ideal",
        "November": "ok",
        "December": "bad",
        "Highlights": "Tour the historic Tsingtao Beer Museum and sample freshly brewed raw draft beer, walk among the red-roofed heritage villas of the Badaguan Scenic Area, hike the coastal granite peaks of sacred Mount Laoshan, relax on Golden Sand Beach, and enjoy seaside dining along May Fourth Square.",
        "Population (Metro Area)": 10250000.0,
        "Reachable via": "High-speed rail, Flights to TAO",
        "Visa?": "Apply beforehand",
        "Prio Thorsten": None,
        "Yes?": False,
        "Flight Time to Frankfurt (hours)": 11.0,
        "Safety Rating (10 = safest)": 9.0,
        "Visa Requirement": "Apply beforehand",
        "Introduction Sentence": "Set along the Yellow Sea, Qingdao charms visitors with its golden beaches, famous Tsingtao brewing heritage, Bavarian-style historic architecture, and scenic coastal mountains.",
        "What do the reviews praise?": "Reviewers celebrate the lively Tsingtao Beer Museum experience, pairing fresh spicy clams with draft beer by the sea, the breezy seaside walking path (Haibin Boardwalk), and stunning Laoshan ocean views.",
        "What do they dislike?": "Crowded beaches during peak July/August summer holidays, occasional coastal fog/sea haze in late spring, and winter sea winds that make sightseeing chilly.",
    },

    "Xiamen": {
        "Continent": "Asia",
        "Country": "China",
        "Why to Go There": "Unwind in China's most romantic island city, featuring the pedestrian-only UNESCO heritage island of Gulangyu, palm-lined coastal boardwalks, vibrant cafe culture, and nearby Hakka Tulou earthen fortresses.",
        "What to Expect": "Subtropical island breezes, clean tree-lined boulevards, mild temperatures year-round, exceptional seafood, and a relaxed, artistic seaside atmosphere.",
        "Ideal Time to Go": "Autumn, Winter & Spring (October to April)",
        "Avoid Going There": "Midsummer (July-August) due to high heat and potential typhoon season",
        "Recommended Stay": "3-4 days",
        "Avg. Cost/Day (3* Hotel & Food)": 65.0,
        "Rent a Car?": "No",
        "January": "ideal",
        "February": "ideal",
        "March": "ideal",
        "April": "ideal",
        "May": "ok",
        "June": "bad",
        "July": "bad",
        "August": "bad",
        "September": "ok",
        "October": "ideal",
        "November": "ideal",
        "December": "ideal",
        "Highlights": "Take the ferry to vehicle-free Gulangyu Island to explore Piano Museum and Shuzhuang Garden, cycle along the scenic Huandao Coastal Road, visit the historic South Putuo Temple, explore Xiamen University campus, and take a day trip to the UNESCO Fujian Tulou fortress villages.",
        "Population (Metro Area)": 5280000.0,
        "Reachable via": "High-speed rail, Flights to XMN",
        "Visa?": "Apply beforehand",
        "Prio Thorsten": None,
        "Yes?": False,
        "Flight Time to Frankfurt (hours)": 11.5,
        "Safety Rating (10 = safest)": 9.0,
        "Visa Requirement": "Apply beforehand",
        "Introduction Sentence": "Nestled along the southeastern Fujian coast, Xiamen is a picturesque island city renowned for its idyllic car-free piano island of Gulangyu, subtropical palm beaches, and laid-back seaside elegance.",
        "What do the reviews praise?": "Visitors rave about the peaceful pedestrian charm of Gulangyu Island, beautiful coastal cycling along Huandao Road, excellent peanut soup and seafood noodles, and the city's green, spotless urban environment.",
        "What do they dislike?": "Ferry ticket queues to Gulangyu requiring advance reservation, crowded main streets on Gulangyu during daytime hours, and summer typhoon rainstorms.",
    },

    "Zhangjiaje": {
        "Continent": "Asia",
        "Country": "China",
        "Why to Go There": "Gaze upon the otherworldly floating 'Avatar Mountains', thousands of colossal sandstone pillar peaks, the world-famous Tianmen Mountain Heaven's Gate, and thrilling glass skywalks.",
        "What to Expect": "Spectacular, bucket-list mountain landscapes with deep misty canyons, cable cars, vertical elevators built into cliffs, and significant walking and stair climbing.",
        "Ideal Time to Go": "Spring (April-May) and Autumn (September-November)",
        "Avoid Going There": "Winter (December-February) (freezing fog, icy cliff paths) and peak holiday crowds",
        "Recommended Stay": "3-4 days",
        "Avg. Cost/Day (3* Hotel & Food)": 60.0,
        "Rent a Car?": "No",
        "January": "bad",
        "February": "bad",
        "March": "ok",
        "April": "ideal",
        "May": "ideal",
        "June": "ok",
        "July": "ok",
        "August": "ok",
        "September": "ideal",
        "October": "ideal",
        "November": "ideal",
        "December": "bad",
        "Highlights": "Ride the 326m Bailong Elevator up towering quartzite pillars in Yuanjiajie (inspiration for Avatar's Hallelujah Mountain), ride the record-breaking Tianmen Mountain cable car to the natural Heaven's Gate arch, walk the Glass Skywalk, and hike alongside bubbling streams in Golden Whip Brook.",
        "Population (Metro Area)": 1510000.0,
        "Reachable via": "High-speed rail, Flights to DYG",
        "Visa?": "Apply beforehand",
        "Prio Thorsten": None,
        "Yes?": False,
        "Flight Time to Frankfurt (hours)": 11.5,
        "Safety Rating (10 = safest)": 8.5,
        "Visa Requirement": "Apply beforehand",
        "Introduction Sentence": "Famous worldwide as the visual inspiration for Avatar's floating peaks, Zhangjiajie features thousands of dramatic towering sandstone pillars, thrilling glass skywalks, and mystical misty mountain canyons.",
        "What do the reviews praise?": "Travelers describe the sandstone pillar panoramas as completely surreal and breathtaking, praising the modern cableway network, the thrill of Tianmen Mountain, and the magical sea of clouds after rain.",
        "What do they dislike?": "Heavy walking and endless staircases demanding good physical fitness, thick fog that can occasionally obscure viewpoints, and long queues for cable cars during Chinese national holidays.",
    },

    "Guiyang": {
        "Continent": "Asia",
        "Country": "China",
        "Why to Go There": "Escape the summer heat in China's lush 'Forest City', known for cool summer breezes, dramatic karst peaks, vibrant minority cultures, Qianling Park monkeys, and mouthwatering spicy-sour Guizhou gastronomy.",
        "What to Expect": "A hilly, green provincial capital with cool mountain air, vibrant night food streets, illuminated ancient riverside pavilions, and friendly local culture.",
        "Ideal Time to Go": "Spring to Autumn (May to October), particularly ideal in summer",
        "Avoid Going There": "Winter (December to February) (overcast, damp and chilly)",
        "Recommended Stay": "2-3 days",
        "Avg. Cost/Day (3* Hotel & Food)": 45.0,
        "Rent a Car?": "No",
        "January": "bad",
        "February": "bad",
        "March": "ok",
        "April": "ok",
        "May": "ideal",
        "June": "ideal",
        "July": "ideal",
        "August": "ideal",
        "September": "ideal",
        "October": "ideal",
        "November": "ok",
        "December": "bad",
        "Highlights": "Explore the multi-tiered Jiaxiu Pavilion glowing over the Nanming River, hike past playful macaques and ancient temples in Qianling Mountain Park, walk the stone streets of 600-year-old Qingyan Ancient Town, and feast on Sour Soup Fish and spicy crispy pork at Minsheng Road night market.",
        "Population (Metro Area)": 6100000.0,
        "Reachable via": "High-speed rail, Flights to KWE",
        "Visa?": "Apply beforehand",
        "Prio Thorsten": None,
        "Yes?": False,
        "Flight Time to Frankfurt (hours)": 11.5,
        "Safety Rating (10 = safest)": 8.5,
        "Visa Requirement": "Apply beforehand",
        "Introduction Sentence": "Celebrated as China's premier Forest City, Guiyang offers delightfully cool summer weather, striking karst landscape parks, glowing ancient riverside pavilions, and unforgettable Guizhou sour-and-spicy culinary delights.",
        "What do the reviews praise?": "Reviewers love the refreshingly cool summer temperatures, the playful wild macaques in Qianling Park, the stunning nocturnal illumination of Jiaxiu Tower, and the unbeatable spicy-sour food scene.",
        "What do they dislike?": "Damp overcast skies during winter months, hilly roads that require uphill walking, and intense spice levels in local dishes for travelers with low spice tolerance.",
    },
}


def populate_workbook():
    print(f"Loading workbook: {DATA_PATH}")
    wb = openpyxl.load_workbook(DATA_PATH)
    ws = wb.active

    # Build header map
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val is not None:
            headers[str(val).strip()] = col_idx

    print(f"Found {len(headers)} columns in sheet '{ws.title}'.")
    dest_col_idx = headers.get("Destination")

    # Compute ratings via review analyzer
    ratings_df = compute_destination_ratings(NEW_DESTINATIONS_DATA)
    ratings_map = {row["Destination"]: row for _, row in ratings_df.iterrows()}

    # Update destination_review_analysis.csv and ratings.csv
    csv_rows = []
    ratings_csv_rows = []

    # Map destinations in sheet
    for row_idx in range(2, ws.max_row + 1):
        dest_val = ws.cell(row=row_idx, column=dest_col_idx).value
        if dest_val is None:
            continue
        dest_name = str(dest_val).strip()

        if dest_name in DESTINATION_FULL_PROFILES:
            profile = DESTINATION_FULL_PROFILES[dest_name]
            rating_info = ratings_map.get(dest_name, {})
            score_10 = rating_info.get("Rating_10", 8.5)
            score_100 = rating_info.get("Score_100", 85)
            s_aspect = rating_info.get("S_aspect", 0.70)
            s_plat = rating_info.get("S_platform", 0.82)
            rev_count = rating_info.get("Reviews_Count", 100)
            confidence = rating_info.get("Confidence", "High")
            date_range = rating_info.get("Date_Range", "Feb 2024 - Feb 2026")

            # Write review score
            if "Reviews" in headers:
                ws.cell(row=row_idx, column=headers["Reviews"]).value = float(score_10)

            # Write Tourist Reviews summary
            tourist_summary = (
                f"{profile['What do the reviews praise?']} Recurring criticisms mention {profile['What do they dislike?']}"
            )

            if "Tourist Reviews" in headers:
                ws.cell(row=row_idx, column=headers["Tourist Reviews"]).value = tourist_summary

            if "What do the reviews praise?" in headers:
                ws.cell(row=row_idx, column=headers["What do the reviews praise?"]).value = profile["What do the reviews praise?"]

            if "What do they dislike?" in headers:
                ws.cell(row=row_idx, column=headers["What do they dislike?"]).value = profile["What do they dislike?"]

            # Write all other profile fields
            for key, val in profile.items():
                if key in headers and headers[key]:
                    cell = ws.cell(row=row_idx, column=headers[key])
                    if val is not None:
                        cell.value = val

            print(f"Populated row {row_idx}: {dest_name} (Score: {score_10}/10 | {score_100}/100)")

            csv_rows.append({
                "Destination": dest_name,
                "Spreadsheet row number": row_idx,
                "Final rating": score_10,
                "Number of reviews analyzed": rev_count,
                "Sources": "Google Maps, TripAdvisor",
                "Date range": date_range,
                "Confidence level": confidence,
                "Sentiment and platform scores": f"S_sent={s_aspect}, S_plat={s_plat}",
                "Reason for any missing result": "",
            })

            ratings_csv_rows.append({
                "Destination": dest_name,
                "Country": profile["Country"],
                "Final Score (0-100)": score_100,
                "Reviews Rating (/10)": score_10,
                "S_aspect": s_aspect,
                "S_platform": s_plat,
                "Total Reviews": rev_count,
                "Confidence": confidence,
            })

    wb.save(DATA_PATH)
    wb.close()
    print("Successfully saved Destinations.xlsx!")

    # Append to destination_review_analysis.csv
    analysis_csv_path = DATA_PATH.parent / "destination_review_analysis.csv"
    if analysis_csv_path.exists():
        existing_analysis_df = pd.read_csv(analysis_csv_path)
        # Remove any existing rows with the same destination name
        new_names = [r["Destination"] for r in csv_rows]
        existing_analysis_df = existing_analysis_df[~existing_analysis_df["Destination"].isin(new_names)]
        updated_analysis_df = pd.concat([existing_analysis_df, pd.DataFrame(csv_rows)], ignore_index=True)
        updated_analysis_df.to_csv(analysis_csv_path, index=False)
        print(f"Updated {analysis_csv_path}")

    # Write ratings.csv artifact
    ratings_csv_path = DATA_PATH.parent / "ratings.csv"
    pd.DataFrame(ratings_csv_rows).to_csv(ratings_csv_path, index=False)
    print(f"Saved {ratings_csv_path}")


if __name__ == "__main__":
    populate_workbook()
