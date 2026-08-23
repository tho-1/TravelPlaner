"""Populate the "Malaria risk?" column in the destination workbook.

Values are normally ``"{status} — {short description}"`` where status is one
of ``no`` / ``near zero`` / ``yes``.

Special rule requested for EUR:
    If the row is in EUR, write exactly ``No`` with no additional comment.

City-first behaviour:
1. If the row's region is EUR, write ``No``.
2. Otherwise use a curated destination/city override.
3. Otherwise fall back to country data only when the country is malaria-free.
4. For malaria-present countries without a city override, skip with a warning
   instead of over-generalising country data to the city.

Usage:
    python populate_malaria_risk.py                  # fill every resolvable row
    python populate_malaria_risk.py "Chennai"        # fill a single destination

Raises ``WorkbookLockedError`` if the file is locked by another program.
"""

from __future__ import annotations

import sys
import unicodedata
from pathlib import Path
from typing import Dict, Optional, Tuple

import openpyxl

from data_utils import DATA_PATH, WorkbookLockedError, _find_destination_sheet, load_destinations
from deepseek_client import generate_malaria_risk

COLUMN = "Malaria risk?"

# Exact value requested for EUR rows.
EUR_VALUE = "No"

# Header names that may contain the regional code.
# Extend this if your workbook uses another column name.
REGION_HEADER_CANDIDATES = (
    "Region",
    "Region Code",
    "WHO Region",
    "World Region",
    "Travel Region",
    "Country Group",
    "Group",
    "Area",
    "Continent",
)

# Values treated as EUR. Normalization removes spaces, punctuation and accents.
_EUR_REGION_VALUES = {
    "eur",
    "europe",
    "eu",
    "euro",
    "european",
    "europeanunion",
    "europeanregion",
    "eurregion",
    "regioneur",
    "whoeur",
    "whoeuro",
    "whoeurope",
}

# Fallback used only when no region value is present.
# If your workbook has a reliable Region column, this is mostly a safety net.
_EUR_COUNTRY_FALLBACK = {
    "albania",
    "andorra",
    "armenia",
    "austria",
    "belarus",
    "belgium",
    "bosnia and herzegovina",
    "bulgaria",
    "croatia",
    "cyprus",
    "czech republic",
    "czechia",
    "denmark",
    "estonia",
    "finland",
    "france",
    "germany",
    "greece",
    "hungary",
    "iceland",
    "ireland",
    "italy",
    "kosovo",
    "latvia",
    "liechtenstein",
    "lithuania",
    "luxembourg",
    "malta",
    "moldova",
    "monaco",
    "montenegro",
    "netherlands",
    "north macedonia",
    "norway",
    "poland",
    "portugal",
    "romania",
    "russia",
    "san marino",
    "serbia",
    "slovakia",
    "slovenia",
    "spain",
    "sweden",
    "switzerland",
    "turkey",
    "turkiye",
    "ukraine",
    "united kingdom",
    "uk",
    "great britain",
    "england",
    "scotland",
    "wales",
    "northern ireland",
    "vatican city",
}


def _normalize(value: object) -> str:
    """Normalize text for robust matching.

    Examples:
        "Chennai"       -> "chennai"
        "Bogotá"        -> "bogota"
        "Cape Town"     -> "capetown"
        "U.S.A."        -> "usa"
    """
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(
        ch
        for ch in text
        if not unicodedata.combining(ch) and ch.isalnum()
    )


# Malaria-free countries/territories.
#
# Country fallback is allowed only for these entries because a certified
# malaria-free country implies that any city in that country is also "no".
#
# Do NOT add malaria-present or near-zero countries here unless you want
# country-level answers again.
MALARIA_FREE_BY_COUNTRY: Dict[str, Tuple[str, str]] = {
    # ── No malaria / no local transmission ───────────────────────────────
    "algeria": ("no", "no local transmission"),
    "australia": ("no", "no local transmission"),
    "austria": ("no", "no local transmission"),
    "bosnia and herzegovina": ("no", "no local transmission"),
    "canada": ("no", "no local transmission"),
    "chile": ("no", "no local transmission"),
    "china": ("no", "no local transmission"),
    "costa rica": ("no", "no local transmission"),
    "cyprus": ("no", "no local transmission"),
    "estonia": ("no", "no local transmission"),
    "finland": ("no", "no local transmission"),
    "france": ("no", "no local transmission"),
    "greece": ("no", "no local transmission"),
    "iceland": ("no", "no local transmission"),
    "japan": ("no", "no local transmission"),
    "kyrgyzstan": ("no", "no local transmission"),
    "mongolia": ("no", "no local transmission"),
    "new zealand": ("no", "no local transmission"),
    "norway": ("no", "no local transmission"),
    "paraguay": ("no", "no local transmission"),
    "poland": ("no", "no local transmission"),
    "puerto rico": ("no", "no local transmission"),
    "qatar": ("no", "no local transmission"),
    "romania": ("no", "no local transmission"),
    "singapore": ("no", "no local transmission"),
    "spain": ("no", "no local transmission"),
    "sri lanka": ("no", "no local transmission"),
    "turkey": ("no", "no local transmission"),
    "turkiye": ("no", "no local transmission"),
    "uae": ("no", "no local transmission"),
    "united arab emirates": ("no", "no local transmission"),
    "united kingdom": ("no", "no local transmission"),
    "uk": ("no", "no local transmission"),
    "united states": ("no", "no local transmission"),
    "united states of america": ("no", "no local transmission"),
    "usa": ("no", "no local transmission"),
    "us": ("no", "no local transmission"),
    "uruguay": ("no", "no local transmission"),
    "uzbekistan": ("no", "no local transmission"),
}


# City/destination-level overrides.
#
# This is the table you should maintain for destinations in malaria-present
# countries. Unknown cities in malaria-present countries are intentionally
# skipped with a warning.
#
# Chennai example:
#   "near zero" is used here because the requested wording wants to say that
#   the city itself is not the main problem, but trips outside the city are.
#   If your official source says Chennai has meaningful urban transmission,
#   change the status to "yes".
MALARIA_BY_DESTINATION: Dict[str, Tuple[str, str]] = {
    # India
    "Chennai": ("near zero", "minimal city risk; risk outside city"),
    "Madras": ("near zero", "minimal city risk; risk outside city"),

    # Indonesia — common tourist cities in Java/Bali.
    "Jakarta": ("near zero", "no local transmission in Java"),
    "Yogyakarta": ("near zero", "no local transmission in Java"),
    "Denpasar": ("near zero", "no local transmission in Bali"),
    "Ubud": ("near zero", "no local transmission in Bali"),

    # High-altitude Andean cities.
    "Bogotá": ("no", "high altitude, no transmission"),
    "Medellín": ("near zero", "minimal risk at this altitude"),
    "Lima": ("no", "no transmission in coastal city"),
    "Cusco": ("no", "high altitude, no transmission"),
    "Arequipa": ("no", "high altitude, no transmission"),
    "Quito": ("no", "high altitude, no transmission"),
    "La Paz": ("no", "high altitude, no transmission"),
    "Sucre": ("no", "high altitude, no transmission"),

    # Cities in risk countries where the city itself is effectively safe.
    "Panama City": ("near zero", "no risk in city; risk outside"),
    "Salvador": ("near zero", "no risk in coastal city"),
    "Kathmandu": ("near zero", "no risk in valley; risk outside"),
    "Windhoek": ("near zero", "no risk in city; risk outside"),
    "Johannesburg": ("near zero", "no risk in city; risk outside"),
    "Cape Town": ("near zero", "no risk in city; risk outside"),

    # Additional examples. Verify against current WHO/CDC guidance.
    "Bangkok": ("near zero", "minimal city risk; risk outside city"),
    "Kuala Lumpur": ("near zero", "no risk in city; risk outside"),
    "Mexico City": ("no", "high altitude, no transmission"),
}


# Normalized lookup dictionaries.
EUR_REGION_VALUES = {_normalize(x) for x in _EUR_REGION_VALUES}
EUR_COUNTRY_FALLBACK = {_normalize(x) for x in _EUR_COUNTRY_FALLBACK}
MALARIA_FREE_BY_COUNTRY_NORM = {
    _normalize(k): v for k, v in MALARIA_FREE_BY_COUNTRY.items()
}
MALARIA_BY_DESTINATION_NORM = {
    _normalize(k): v for k, v in MALARIA_BY_DESTINATION.items()
}


def _is_eur(region: object, country: object) -> bool:
    """Return True if the row should be treated as EUR.

    If a non-empty region value exists, it is authoritative. This avoids
    treating overseas territories as EUR merely because the country column
    says, for example, "France" or "Netherlands".

    If the region value is empty, fall back to the European country list.
    """
    region_key = _normalize(region)

    if region_key:
        if region_key in EUR_REGION_VALUES:
            return True
        # Handles values like "Europe (Western)" or "Europe EUR".
        if "europe" in region_key:
            return True
        # A region exists but is not EUR; do not infer EUR from country.
        return False

    country_key = _normalize(country)
    return country_key in EUR_COUNTRY_FALLBACK


def resolve_malaria(destination: str, country: str) -> Optional[Tuple[str, str]]:
    """Return ``(status, description)`` for a destination, or ``None``.

    Resolution order:
    1. City/destination override.
    2. Country fallback, but only if the country is malaria-free.

    Unknown cities in malaria-present countries return ``None`` so that the
    caller can emit a warning instead of applying country-level risk.
    """
    dest_key = _normalize(destination)
    country_key = _normalize(country)

    override = MALARIA_BY_DESTINATION_NORM.get(dest_key)
    if override is not None:
        return override

    country_entry = MALARIA_FREE_BY_COUNTRY_NORM.get(country_key)
    if country_entry is not None and country_entry[0] == "no":
        return country_entry

    return None


def populate_malaria_risk(
    destination: Optional[str] = None,
    country: Optional[str] = None,
    path: Path = DATA_PATH,
) -> Tuple[int, list]:
    """Fill the 'Malaria risk?' column and save once.

    With ``destination``, fills only that row. Without it, fills every
    resolvable row.

    Returns ``(filled_count, warnings)`` and raises ``WorkbookLockedError``
    if the workbook is locked.
    """
    sheet_name = _find_destination_sheet(path)
    if sheet_name is None:
        return 0, ["Could not find destination sheet in workbook."]

    try:
        wb = openpyxl.load_workbook(path)
    except PermissionError as exc:
        raise WorkbookLockedError(
            "Destinations.xlsx is currently open in another program (e.g. Excel). "
            "Please close the file and press Retry."
        ) from exc

    ws = wb[sheet_name]

    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val is not None:
            headers[str(val).strip()] = col_idx

    headers_norm = {_normalize(name): idx for name, idx in headers.items()}

    def find_col(*names: str) -> Optional[int]:
        for name in names:
            idx = headers_norm.get(_normalize(name))
            if idx is not None:
                return idx
        return None

    dest_col = find_col(
        "Destination",
        "City",
        "Place",
        "Location",
        "Destination / City",
    )
    country_col = find_col(
        "Country",
        "Country Name",
        "Country / Region",
        "Nation",
    )
    region_col = find_col(*REGION_HEADER_CANDIDATES)

    if dest_col is None:
        wb.close()
        return 0, ["Could not find the 'Destination' column."]

    malaria_col = find_col(COLUMN)
    if malaria_col is None:
        malaria_col = ws.max_column + 1
        ws.cell(row=1, column=malaria_col, value=COLUMN)
        headers_norm[_normalize(COLUMN)] = malaria_col

    warnings = []
    filled = 0

    def _write(row_idx: int, dest: str) -> None:
        nonlocal filled

        workbook_country = ""
        if country_col:
            raw = ws.cell(row=row_idx, column=country_col).value
            workbook_country = str(raw).strip() if raw is not None else ""

        # Explicit country passed via API/CLI wins over workbook country.
        eff_country = (country or "").strip() or workbook_country

        region_value = None
        if region_col:
            region_value = ws.cell(row=row_idx, column=region_col).value

        # Requested EUR rule: exact "No", no comment.
        if _is_eur(region_value, eff_country):
            ws.cell(row=row_idx, column=malaria_col).value = EUR_VALUE
            filled += 1
            return

        result = resolve_malaria(dest, eff_country)
        if result is None:
            try:
                ai_result = generate_malaria_risk(dest, eff_country)
            except RuntimeError as exc:
                warnings.append(f"{dest}: DeepSeek lookup failed: {exc}")
                return
            result = (ai_result["status"], ai_result["description"])

        status, description = result
        ws.cell(row=row_idx, column=malaria_col).value = f"{status} — {description}"
        filled += 1

    if destination is not None:
        dest_key = _normalize(destination)
        target_row = None
        target_value = None

        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=dest_col).value
            if val is not None and _normalize(val) == dest_key:
                target_row = row_idx
                target_value = str(val).strip()
                break

        if target_row is None:
            wb.close()
            return 0, [f"Destination '{destination}' was not found in the workbook."]

        _write(target_row, target_value)
    else:
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=dest_col).value
            if val is not None and str(val).strip():
                _write(row_idx, str(val).strip())

    try:
        wb.save(path)
    except PermissionError as exc:
        wb.close()
        raise WorkbookLockedError(
            "Destinations.xlsx is currently open in another program (e.g. Excel). "
            "Please close the file and press Retry."
        ) from exc

    wb.close()

    # Invalidate cached destination data if data_utils uses lru_cache or similar.
    from data_utils import _clear_destination_cache
    _clear_destination_cache()

    return filled, warnings


if __name__ == "__main__":
    target = sys.argv[1].strip() if len(sys.argv) > 1 else None

    try:
        n, warns = populate_malaria_risk(destination=target)
    except WorkbookLockedError as exc:
        print(f"LOCKED: {exc}")
        sys.exit(1)

    print(f"Filled {n} row(s).")
    for w in warns:
        print(f"  !! {w}")