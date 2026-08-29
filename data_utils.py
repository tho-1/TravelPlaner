from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Optional, Tuple

import openpyxl
import pandas as pd
import streamlit as st


DATA_PATH = Path(__file__).resolve().parent / "Destinations.xlsx"

OPEN_TABS_PATH = Path(__file__).resolve().parent / "open_destinations.json"


def load_open_destinations() -> list:
    """Load the persisted list of open destination tabs (JSON file)."""
    try:
        if OPEN_TABS_PATH.exists():
            data = json.loads(OPEN_TABS_PATH.read_text(encoding="utf-8"))
            if isinstance(data, list):
                return [str(value) for value in data if str(value).strip()]
    except Exception:
        pass
    return []


def save_open_destinations(destinations) -> None:
    """Persist the list of open destination tabs to a JSON file (best effort)."""
    try:
        OPEN_TABS_PATH.write_text(
            json.dumps([str(d) for d in destinations], ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception:
        pass


def to_be_researched_mask(frame: pd.DataFrame, column: str) -> pd.Series:
    """Return a boolean mask of rows whose 'to be researched' column is truthy.

    Mirrors the value parsing used on the destination detail page: True / 1 /
    "x" / "yes" / "y" / "ja" / "j" count as requiring research, and any string
    containing an 'x' does too (the app's convention).
    """
    mask = pd.Series(False, index=frame.index, dtype=bool)
    values = frame[column].tolist()
    for pos, value in enumerate(values):
        if pd.isna(value):
            continue
        if isinstance(value, bool):
            mask.iloc[pos] = value
        elif isinstance(value, (int, float)):
            mask.iloc[pos] = (value == 1)
        else:
            text = str(value).strip().lower()
            mask.iloc[pos] = text in {"yes", "y", "true", "1", "1.0", "ja", "j", "x"} or "x" in text
    return mask


class WorkbookLockedError(RuntimeError):
    """Raised when Destinations.xlsx cannot be written because another program
    (e.g. Excel, OneDrive) currently holds an exclusive lock on the file."""


def normalize_text(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def find_column(columns, aliases) -> Optional[str]:
    normalized_aliases = [normalize_text(alias) for alias in aliases]
    for column in columns:
        if normalize_text(column) in normalized_aliases:
            return column
    for column in columns:
        normalized = normalize_text(column)
        if any(alias in normalized for alias in normalized_aliases):
            return column
    for column in columns:
        normalized = normalize_text(column)
        for alias in normalized_aliases:
            if alias in normalized or normalized in alias:
                return column
    return None


def parse_numeric(series: pd.Series) -> pd.Series:
    cleaned = series.astype(str).str.replace(r"[^0-9,.-]", "", regex=True)
    cleaned = cleaned.replace("", pd.NA)
    cleaned = cleaned.str.replace(",", "", regex=False)
    return pd.to_numeric(cleaned, errors="coerce")


def classify_bool(series: pd.Series) -> pd.Series:
    result = []
    for value in series:
        if pd.isna(value):
            result.append(pd.NA)
            continue
        if isinstance(value, bool):
            result.append(value)
            continue
        if isinstance(value, (int, float)):
            if value == 1:
                result.append(True)
                continue
            elif value == 0:
                result.append(False)
                continue
        text = str(value).strip().lower().rstrip(",.").strip()
        if text in {"yes", "y", "true", "1", "1.0", "ja", "j", "x"} or "x" in text:
            result.append(True)
        elif text in {"no", "n", "false", "0", "0.0", "nein"}:
            result.append(False)
        else:
            result.append(pd.NA)
    return pd.Series(result, dtype="object")


def discover_destination_sheet(path: Path) -> pd.DataFrame:
    with pd.ExcelFile(path, engine="openpyxl") as excel_file:
        for sheet_name in excel_file.sheet_names:
            try:
                df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
            except Exception:
                continue
            if df is None or df.empty:
                continue
            columns = [str(col) for col in df.columns]
            if not columns:
                continue
            if any(
                keyword in " ".join(columns).lower()
                for keyword in ["destination", "continent", "safety", "cost", "flight", "population", "month"]
            ):
                return df
    return pd.read_excel(path, engine="openpyxl")


@st.cache_data(show_spinner=False)
def _load_destinations_cached(path: Path, modified_ns: int) -> Tuple[pd.DataFrame, dict]:
    df = discover_destination_sheet(path)
    df = df.copy()
    df.columns = [str(col) for col in df.columns]

    destination_col = find_column(df.columns, ["destination", "destinations", "city", "place", "name"])
    country_col = find_column(df.columns, ["country", "land", "nation", "state"])
    continent_col = find_column(df.columns, ["continent", "region"])
    eu_col = find_column(df.columns, ["eu", "eu?", "europeanunion", "in europe", "yes", "yes?"])
    if eu_col is None:
        # Some workbooks use a column header like "Yes?" for EU membership.
        for col in df.columns:
            if normalize_text(str(col)) in {"yes", "yes?", "yesquestion"}:
                eu_col = col
                break
    nearer_col = find_column(df.columns, ["näherer", "auswahl", "nahe", "nearer", "selection"])
    visited_col = find_column(df.columns, ["visited", "visited?"])
    to_be_researched_col = find_column(df.columns, ["to be researched", "toberesearched", "research needed", "researched?"])
    safety_col = find_column(df.columns, ["safety", "safetyrating", "security", "risk"])
    cost_col = find_column(df.columns, ["costday", "costperday", "dailycost", "cost", "costday"])
    flight_col = find_column(df.columns, ["flighttime", "flight", "tofra", "travel", "duration"])
    population_col = find_column(df.columns, ["population", "inhabitants", "inhabitant"])
    reviews_col = find_column(df.columns, ["reviews", "reviewcount", "review score"])
    prio_col = find_column(df.columns, ["prio thorsten", "prio", "thorsten"])
    visa_requirement_col = find_column(df.columns, ["visa requirement", "visarequirement", "visareq"])
    highlights_col = find_column(df.columns, ["highlights", "whygo", "notes", "pros", "bestfor"])
    intro_col = find_column(df.columns, ["introduction sentence", "introduction", "intro sentence", "intro"])
    why_col = find_column(df.columns, ["why to go there", "why go there", "whytogothere", "whygo"])
    expect_col = find_column(df.columns, ["what to expect", "whattoexpect", "expectations", "expect"])
    tourist_reviews_col = find_column(df.columns, ["tourist reviews", "tourist review", "touristreviews", "touristreview"])
    status_col = find_column(df.columns, ["data status", "status"])
    comment_col = find_column(df.columns, ["comment", "kommentar", "anmerkung"])
    malaria_risk_col = find_column(df.columns, ["malaria risk", "malaria", "malariarisk"])
    food_spiciness_col = find_column(df.columns, ["food - spicyness", "food spicyness"])
    food_description_col = find_column(df.columns, ["food - description", "food description"])

    if destination_col is None:
        destination_col = df.columns[0]
    if continent_col is None:
        continent_col = df.columns[1] if len(df.columns) > 1 else destination_col

    if cost_col:
        df[cost_col] = parse_numeric(df[cost_col])
    if flight_col:
        df[flight_col] = parse_numeric(df[flight_col])
    if safety_col:
        df[safety_col] = parse_numeric(df[safety_col])
    if population_col:
        df[population_col] = parse_numeric(df[population_col])
    if eu_col:
        df[eu_col] = classify_bool(df[eu_col])
    if nearer_col:
        df[nearer_col] = classify_bool(df[nearer_col])
    if visited_col:
        df[visited_col] = classify_bool(df[visited_col])
    if to_be_researched_col:
        df[to_be_researched_col] = classify_bool(df[to_be_researched_col])

    _month_names = [
        "january", "february", "march", "april", "may", "june",
        "july", "august", "september", "october", "november", "december",
        "jan", "feb", "mar", "apr", "jun", "jul", "aug", "sep", "oct", "nov", "dec"
    ]
    _climate_tokens = ["high", "low", "rain", "aqi", "temp", "day", "days", "cost", "prio", "mm"]

    month_columns = [
        col for col in df.columns
        if any(normalize_text(col) == m or normalize_text(col).startswith(m) for m in _month_names)
        and not any(token in normalize_text(col) for token in _climate_tokens)
    ]
    metadata = {
        "destination_col": destination_col,
        "country_col": country_col,
        "continent_col": continent_col,
        "eu_col": eu_col,
        "nearer_col": nearer_col,
        "visited_col": visited_col,
        "to_be_researched_col": to_be_researched_col,
        "safety_col": safety_col,
        "cost_col": cost_col,
        "flight_col": flight_col,
        "population_col": population_col,
        "reviews_col": reviews_col,
        "prio_col": prio_col,
        "visa_requirement_col": visa_requirement_col,
        "highlights_col": highlights_col,
        "intro_col": intro_col,
        "why_col": why_col,
        "expect_col": expect_col,
        "tourist_reviews_col": tourist_reviews_col,
        "status_col": status_col,
        "comment_col": comment_col,
        "malaria_risk_col": malaria_risk_col,
        "food_spiciness_col": food_spiciness_col,
        "food_description_col": food_description_col,
        "month_columns": month_columns,
    }
    return df, metadata


def load_destinations(path: Path = DATA_PATH) -> Tuple[pd.DataFrame, dict]:
    """Load destinations and refresh automatically when the workbook changes."""
    modified_ns = path.stat().st_mtime_ns
    return _load_destinations_cached(path, modified_ns)


def _clear_destination_cache() -> None:
    _load_destinations_cached.clear()


def _find_destination_sheet(path: Path) -> Optional[str]:
    """Return the name of the sheet that contains the destination data."""
    with pd.ExcelFile(path, engine="openpyxl") as excel_file:
        for sheet_name in excel_file.sheet_names:
            try:
                df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl", nrows=0)
            except Exception:
                continue
            columns = [str(col) for col in df.columns]
            if any(
                keyword in " ".join(columns).lower()
                for keyword in ["destination", "continent", "safety", "cost", "flight", "population", "month"]
            ):
                return sheet_name
    return None


def update_favorite_status(destination_name: str, add: bool, path: Path = DATA_PATH) -> None:
    """Set or clear the 'In nähererer Auswahl 2025?' cell for a destination.

    Uses openpyxl to modify the cell in-place so that formatting,
    formulas, and other sheets in the workbook are preserved.
    """
    sheet_name = _find_destination_sheet(path)
    if sheet_name is None:
        return

    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]

    # Build a header map: column name -> 1-based column index
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col_idx).value
        if cell_value is not None:
            headers[str(cell_value)] = col_idx

    # Find the destination column and nearer column by fuzzy matching
    dest_col_idx = None
    nearer_col_idx = None
    for header_name, col_idx in headers.items():
        lower = header_name.lower()
        if dest_col_idx is None and "destination" in lower:
            dest_col_idx = col_idx
        if nearer_col_idx is None and "auswahl" in lower:
            nearer_col_idx = col_idx

    if dest_col_idx is None or nearer_col_idx is None:
        wb.close()
        return

    # Find the row for this destination
    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=dest_col_idx).value
        if cell_value is not None and str(cell_value).strip().lower() == destination_name.strip().lower():
            target_row = row_idx
            break

    if target_row is None:
        wb.close()
        return

    # Write "x" or clear the cell
    ws.cell(row=target_row, column=nearer_col_idx).value = "x" if add else None
    wb.save(path)
    wb.close()

    # Clear the cached data so the app picks up the change
    _clear_destination_cache()


def update_visited_status(destination_name: str, visited: bool, path: Path = DATA_PATH) -> None:
    """Set the ``Visited?`` value for a destination in the workbook."""
    sheet_name = _find_destination_sheet(path)
    if sheet_name is None:
        return

    try:
        wb = openpyxl.load_workbook(path)
    except PermissionError as exc:
        raise WorkbookLockedError(
            "Destinations.xlsx is currently open in another program (e.g. Excel). "
            "Please close the file and press Retry."
        ) from exc
    ws = wb[sheet_name]

    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col_idx).value
        if cell_value is not None:
            headers[str(cell_value).strip()] = col_idx

    dest_col_idx = None
    visited_col_idx = None
    for header_name, col_idx in headers.items():
        lower = header_name.lower()
        if dest_col_idx is None and "destination" in lower:
            dest_col_idx = col_idx
        if visited_col_idx is None and lower in {"visited", "visited?"}:
            visited_col_idx = col_idx

    if dest_col_idx is None or visited_col_idx is None:
        wb.close()
        return

    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=dest_col_idx).value
        if cell_value is not None and str(cell_value).strip().lower() == destination_name.strip().lower():
            target_row = row_idx
            break

    if target_row is None:
        wb.close()
        return

    ws.cell(row=target_row, column=visited_col_idx).value = bool(visited)
    try:
        wb.save(path)
    except PermissionError as exc:
        wb.close()
        raise WorkbookLockedError(
            "Destinations.xlsx is currently open in another program (e.g. Excel). "
            "Please close the file and press Retry."
        ) from exc
    wb.close()

    _clear_destination_cache()


def update_to_be_researched_status(destination_name: str, to_be_researched: bool, path: Path = DATA_PATH) -> None:
    """Set the ``To be researched`` value for a destination in the workbook."""
    sheet_name = _find_destination_sheet(path)
    if sheet_name is None:
        return

    try:
        wb = openpyxl.load_workbook(path)
    except PermissionError as exc:
        raise WorkbookLockedError(
            "Destinations.xlsx is currently open in another program (e.g. Excel). "
            "Please close the file and press Retry."
        ) from exc
    ws = wb[sheet_name]

    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col_idx).value
        if cell_value is not None:
            headers[str(cell_value).strip()] = col_idx

    dest_col_idx = None
    research_col_idx = None
    for header_name, col_idx in headers.items():
        lower = header_name.lower()
        if dest_col_idx is None and "destination" in lower:
            dest_col_idx = col_idx
        if research_col_idx is None and ("researched" in lower or "research" in lower):
            research_col_idx = col_idx

    if dest_col_idx is None:
        wb.close()
        return

    if research_col_idx is None:
        research_col_idx = ws.max_column + 1
        ws.cell(row=1, column=research_col_idx, value="To be researched")

    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=dest_col_idx).value
        if cell_value is not None and str(cell_value).strip().lower() == destination_name.strip().lower():
            target_row = row_idx
            break

    if target_row is None:
        wb.close()
        return

    ws.cell(row=target_row, column=research_col_idx).value = bool(to_be_researched)
    try:
        wb.save(path)
    except PermissionError as exc:
        wb.close()
        raise WorkbookLockedError(
            "Destinations.xlsx is currently open in another program (e.g. Excel). "
            "Please close the file and press Retry."
        ) from exc
    wb.close()

    _clear_destination_cache()


def update_prio_thorsten(destination_name: str, value: int, path: Path = DATA_PATH) -> None:
    """Update the Prio Thorsten value for a destination in the workbook."""
    sheet_name = _find_destination_sheet(path)
    if sheet_name is None:
        return

    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]

    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col_idx).value
        if cell_value is not None:
            headers[str(cell_value)] = col_idx

    dest_col_idx = None
    prio_col_idx = None
    for header_name, col_idx in headers.items():
        lower = header_name.strip().lower()
        if dest_col_idx is None and "destination" in lower:
            dest_col_idx = col_idx
        if prio_col_idx is None and ("prio" in lower and "thorsten" in lower or lower == "prio thorsten" or "thorsten" in lower):
            prio_col_idx = col_idx

    if dest_col_idx is None or prio_col_idx is None:
        wb.close()
        return

    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=dest_col_idx).value
        if cell_value is not None and str(cell_value).strip().lower() == destination_name.strip().lower():
            target_row = row_idx
            break

    if target_row is None:
        wb.close()
        return

    if value is None or str(value).strip() == "" or str(value).strip().lower() in {"none", "nan", "null", "—"}:
        ws.cell(row=target_row, column=prio_col_idx).value = None
    else:
        try:
            ws.cell(row=target_row, column=prio_col_idx).value = int(float(value))
        except (ValueError, TypeError):
            ws.cell(row=target_row, column=prio_col_idx).value = None

    wb.save(path)
    wb.close()

    _clear_destination_cache()


def update_comment(destination_name: str, value: str, path: Path = DATA_PATH) -> None:
    """Update the Comment value for a destination in the workbook.

    Creates a "Comment" column if the workbook does not have one yet (the
    header is appended at the end of the sheet). Raises ``WorkbookLockedError``
    if the file is currently locked by another program (e.g. Excel).
    """
    sheet_name = _find_destination_sheet(path)
    if sheet_name is None:
        return

    try:
        wb = openpyxl.load_workbook(path)
    except PermissionError as exc:
        raise WorkbookLockedError(
            "Destinations.xlsx is currently open in another program (e.g. Excel). "
            "Please close the file and press Retry."
        ) from exc
    ws = wb[sheet_name]

    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col_idx).value
        if cell_value is not None:
            headers[str(cell_value).strip()] = col_idx

    dest_col_idx = None
    comment_col_idx = None
    for header_name, col_idx in headers.items():
        lower = header_name.strip().lower()
        if dest_col_idx is None and "destination" in lower:
            dest_col_idx = col_idx
        if comment_col_idx is None and lower in {"comment", "kommentar", "anmerkung"}:
            comment_col_idx = col_idx

    if dest_col_idx is None:
        wb.close()
        return

    if comment_col_idx is None:
        comment_col_idx = ws.max_column + 1
        ws.cell(row=1, column=comment_col_idx, value="Comment")

    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=dest_col_idx).value
        if cell_value is not None and str(cell_value).strip().lower() == destination_name.strip().lower():
            target_row = row_idx
            break

    if target_row is None:
        wb.close()
        return

    new_value = "" if value is None else str(value).strip()
    # NOTE: openpyxl treats cell(..., value=None) as "leave unchanged", so use
    # `.value = None` to clear the cell when the comment is emptied.
    ws.cell(row=target_row, column=comment_col_idx).value = new_value if new_value else None

    try:
        wb.save(path)
    except PermissionError as exc:
        wb.close()
        raise WorkbookLockedError(
            "Destinations.xlsx is currently open in another program (e.g. Excel). "
            "Please close the file and press Retry."
        ) from exc
    wb.close()

    _clear_destination_cache()


def update_reviews(
    destination_name: str,
    review_score: float,
    tourist_reviews: str,
    praise: str,
    dislikes: str,
    path: Path = DATA_PATH,
) -> None:
    """Write only the review-related fields for an existing destination."""
    sheet_name = _find_destination_sheet(path)
    if sheet_name is None:
        return
    try:
        wb = openpyxl.load_workbook(path)
    except PermissionError as exc:
        raise WorkbookLockedError(
            "Destinations.xlsx is currently open in another program (e.g. Excel). "
            "Please close the file and press Retry."
        ) from exc
    ws = wb[sheet_name]
    headers = {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value is not None
    }
    dest_col = next((col for name, col in headers.items() if "destination" in name.lower()), None)
    if dest_col is None:
        wb.close()
        return
    target_row = next(
        (
            row for row in range(2, ws.max_row + 1)
            if ws.cell(row, dest_col).value is not None
            and str(ws.cell(row, dest_col).value).strip().lower() == destination_name.strip().lower()
        ),
        None,
    )
    if target_row is None:
        wb.close()
        return
    values = {
        "Reviews": float(review_score),
        "Tourist Reviews": str(tourist_reviews).strip(),
        "What do the reviews praise?": str(praise).strip(),
        "What do they dislike?": str(dislikes).strip(),
    }
    for name, value in values.items():
        if name in headers:
            ws.cell(target_row, headers[name]).value = value
    try:
        wb.save(path)
    except PermissionError as exc:
        wb.close()
        raise WorkbookLockedError(
            "Destinations.xlsx is currently open in another program (e.g. Excel). "
            "Please close the file and press Retry."
        ) from exc
    wb.close()
    _clear_destination_cache()


def update_food(
    destination_name: str,
    spiciness: float,
    description: str,
    path: Path = DATA_PATH,
) -> None:
    """Write the Food rating and description for a destination."""
    sheet_name = _find_destination_sheet(path)
    if sheet_name is None:
        return

    try:
        wb = openpyxl.load_workbook(path)
    except PermissionError as exc:
        raise WorkbookLockedError(
            "Destinations.xlsx is currently open in another program (e.g. Excel). "
            "Please close the file and press Retry."
        ) from exc
    ws = wb[sheet_name]

    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col_idx).value
        if cell_value is not None:
            headers[str(cell_value).strip()] = col_idx

    dest_col_idx = next(
        (col_idx for name, col_idx in headers.items() if "destination" in name.lower()),
        None,
    )
    spiciness_col_idx = next(
        (col_idx for name, col_idx in headers.items() if normalize_text(name) in {"foodspicyness", "foodspiciness"}),
        None,
    )
    description_col_idx = next(
        (col_idx for name, col_idx in headers.items() if normalize_text(name) == "fooddescription"),
        None,
    )

    if dest_col_idx is None:
        wb.close()
        return
    if spiciness_col_idx is None:
        spiciness_col_idx = ws.max_column + 1
        ws.cell(row=1, column=spiciness_col_idx, value="Food - Spicyness")
    if description_col_idx is None:
        description_col_idx = ws.max_column + 1
        ws.cell(row=1, column=description_col_idx, value="Food - Description")

    target_row = next(
        (
            row_idx
            for row_idx in range(2, ws.max_row + 1)
            if ws.cell(row=row_idx, column=dest_col_idx).value is not None
            and str(ws.cell(row=row_idx, column=dest_col_idx).value).strip().lower()
            == destination_name.strip().lower()
        ),
        None,
    )
    if target_row is None:
        wb.close()
        return

    ws.cell(row=target_row, column=spiciness_col_idx).value = float(spiciness)
    ws.cell(row=target_row, column=description_col_idx).value = str(description).strip()
    try:
        wb.save(path)
    except PermissionError as exc:
        wb.close()
        raise WorkbookLockedError(
            "Destinations.xlsx is currently open in another program (e.g. Excel). "
            "Please close the file and press Retry."
        ) from exc
    wb.close()
    _clear_destination_cache()


def get_selected_destination(df: pd.DataFrame, metadata: dict, fallback: Optional[str] = None) -> Optional[pd.Series]:
    destination_col = metadata["destination_col"]
    selected_name = st.session_state.get("selected_destination")
    if not selected_name:
        selected_name = fallback
    if not selected_name:
        return None
    matches = df[df[destination_col].astype(str).str.strip().str.lower() == str(selected_name).strip().lower()]
    if matches.empty:
        return None
    return matches.iloc[0]


def add_new_destination(
    destination_name: str,
    country: str = "Unknown",
    continent: str = "Unknown",
    path: Path = DATA_PATH
) -> Tuple[bool, str]:
    """
    Append a new destination to the Excel workbook with placeholder values.
    Returns (success: bool, message: str).
    """
    dest_clean = destination_name.strip()
    if not dest_clean:
        return False, "Destination name cannot be empty."

    sheet_name = _find_destination_sheet(path)
    if sheet_name is None:
        return False, "Could not find destination sheet in workbook."

    try:
        wb = openpyxl.load_workbook(path)
    except PermissionError as exc:
        raise WorkbookLockedError(
            f"Destinations.xlsx is currently open in another program (e.g. Excel). "
            f"Please close the file and press Retry. ({exc})"
        ) from exc
    ws = wb[sheet_name]

    # Header lookup
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col_idx).value
        if cell_value is not None:
            headers[str(cell_value)] = col_idx

    # Check for destination column
    dest_col_idx = 1
    for h, col_idx in headers.items():
        if "destination" in h.lower():
            dest_col_idx = col_idx
            break

    # Check for existing destination
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=dest_col_idx).value
        if val is not None and str(val).strip().lower() == dest_clean.lower():
            wb.close()
            return False, f"Destination '{dest_clean}' already exists in the workbook (Row {row_idx})."

    # Determine next empty row
    new_row_idx = ws.max_row + 1

    # Populate only the user-supplied identity fields. Everything else remains
    # blank until it is researched and explicitly updated.
    ws.cell(row=new_row_idx, column=dest_col_idx, value=dest_clean)

    status_col_idx = headers.get("Data Status")
    if status_col_idx is None:
        status_col_idx = ws.max_column + 1
        ws.cell(row=1, column=status_col_idx, value="Data Status")

    for h, col_idx in headers.items():
        h_lower = h.lower()
        if "country" in h_lower:
            ws.cell(row=new_row_idx, column=col_idx, value=country.strip() or "Unknown")
        elif "continent" in h_lower:
            ws.cell(row=new_row_idx, column=col_idx, value=continent.strip() or "Unknown")

    ws.cell(row=new_row_idx, column=status_col_idx, value="PLACEHOLDER - UPDATE REQUIRED")

    try:
        wb.save(path)
    except PermissionError as exc:
        wb.close()
        raise WorkbookLockedError(
            f"Destinations.xlsx is currently open in another program (e.g. Excel). "
            f"Please close the file and press Retry. ({exc})"
        ) from exc
    wb.close()

    # Clear cached dataframe so it immediately reloads all rows on next run
    _clear_destination_cache()
    return True, f"Destination '{dest_clean}' has been successfully added to the catalog!"


def load_airlines_color_map(path: Path = DATA_PATH) -> dict[str, dict[str, any]]:
    """Load the Airlines sheet and return a mapping of normalized_name -> {name, color}."""
    try:
        df_air = pd.read_excel(path, sheet_name="Airlines", engine="openpyxl")
    except Exception:
        return {}

    col_map = {}
    for col in df_air.columns:
        c_low = str(col).strip().lower()
        if "airline" in c_low:
            col_map["airline"] = col
        elif "color" in c_low:
            col_map["color"] = col

    airline_col = col_map.get("airline", "Airline")
    color_col = col_map.get("color", "Color")

    mapping = {}
    if airline_col in df_air.columns:
        for _, row in df_air.iterrows():
            val = row.get(airline_col)
            if pd.notna(val) and str(val).strip():
                name = str(val).strip()
                color_val = row.get(color_col) if color_col in df_air.columns else None
                color = str(color_val).strip().lower() if pd.notna(color_val) and str(color_val).strip() else None
                mapping[normalize_text(name)] = {
                    "name": name,
                    "color": color,
                }
    return mapping


def sync_airlines_to_excel(airlines: list[str], path: Path = DATA_PATH) -> int:
    """Ensure all airlines in the list exist in the 'Airlines' sheet of Destinations.xlsx.
    If not, appends them at the bottom with a blank Color.
    Returns the number of newly added airlines.
    """
    if not airlines:
        return 0

    try:
        wb = openpyxl.load_workbook(path)
    except PermissionError as exc:
        raise WorkbookLockedError(
            f"Destinations.xlsx is currently open in another program. ({exc})"
        ) from exc
    except Exception:
        return 0

    sheet_name = "Airlines"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.cell(row=1, column=1, value="Airline")
        ws.cell(row=1, column=2, value="Color")

    # Find headers
    headers = {}
    for col_idx in range(1, max(ws.max_column, 2) + 1):
        v = ws.cell(row=1, column=col_idx).value
        if v:
            headers[str(v).strip().lower()] = col_idx

    airline_col_idx = headers.get("airline", 1)
    color_col_idx = headers.get("color", 2)

    # Ensure header row has titles if missing
    if ws.cell(row=1, column=airline_col_idx).value is None:
        ws.cell(row=1, column=airline_col_idx, value="Airline")
    if ws.cell(row=1, column=color_col_idx).value is None:
        ws.cell(row=1, column=color_col_idx, value="Color")

    # Read existing airlines
    existing_normalized = set()
    for row_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=airline_col_idx).value
        if cell_val is not None and str(cell_val).strip():
            existing_normalized.add(normalize_text(str(cell_val)))

    added_count = 0
    next_row = ws.max_row + 1

    for a in airlines:
        clean_name = str(a).strip()
        if not clean_name:
            continue
        norm = normalize_text(clean_name)
        if norm not in existing_normalized:
            ws.cell(row=next_row, column=airline_col_idx, value=clean_name)
            ws.cell(row=next_row, column=color_col_idx, value=None)
            existing_normalized.add(norm)
            next_row += 1
            added_count += 1

    if added_count > 0:
        try:
            wb.save(path)
        except PermissionError as exc:
            wb.close()
            raise WorkbookLockedError(
                f"Destinations.xlsx is currently locked by another program. ({exc})"
            ) from exc
    wb.close()
    return added_count

