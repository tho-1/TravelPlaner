"""Populate destination food descriptions and spiciness ratings with DeepSeek.

Usage:
    python populate_food.py Asia
    python populate_food.py "Yogyakarta"
"""

from __future__ import annotations

import sys
from typing import Optional

import openpyxl

from data_utils import DATA_PATH, WorkbookLockedError, _find_destination_sheet, load_destinations, update_food
from deepseek_client import generate_food_profile


def populate_food(
    continent: Optional[str] = None,
    destination: Optional[str] = None,
    path=DATA_PATH,
) -> tuple[int, list[str]]:
    """Populate blank Food fields for destinations matching the requested scope."""
    sheet_name = _find_destination_sheet(path)
    if sheet_name is None:
        return 0, ["Could not find destination sheet in workbook."]

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    headers = {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value is not None
    }
    destination_col = headers.get("Destination")
    country_col = headers.get("Country")
    continent_col = headers.get("Continent")
    if destination_col is None or country_col is None:
        wb.close()
        return 0, ["Could not find Destination and Country columns."]

    targets = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, destination_col).value
        country = ws.cell(row, country_col).value
        row_continent = ws.cell(row, continent_col).value if continent_col else None
        if not name or not country:
            continue
        if destination and str(name).strip().lower() != destination.strip().lower():
            continue
        if continent and str(row_continent).strip().lower() != continent.strip().lower():
            continue
        targets.append((str(name).strip(), str(country).strip()))
    wb.close()

    done = 0
    warnings = []
    for name, country in targets:
        try:
            profile = generate_food_profile(name, country)
            update_food(name, profile["spiciness"], profile["description"], path=path)
            done += 1
            print(f"{name}: spiciness {profile['spiciness']:g}/10")
        except (RuntimeError, WorkbookLockedError) as exc:
            warnings.append(f"{name}: {exc}")
            print(f"{name}: ERROR {exc}")
    return done, warnings


if __name__ == "__main__":
    value = sys.argv[1].strip() if len(sys.argv) > 1 else "Asia"
    count, errors = populate_food(
        continent=value if value.lower() in {"asia", "europe", "africa", "oceania", "north america", "south america"} else None,
        destination=None if value.lower() in {"asia", "europe", "africa", "oceania", "north america", "south america"} else value,
    )
    print(f"SUMMARY populated={count} errors={len(errors)}")